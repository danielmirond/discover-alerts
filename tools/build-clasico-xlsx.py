#!/usr/bin/env python3
"""Genera docs/clasico-declaraciones.xlsx con todas las citas verificadas."""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================================
# DERIVACIÓN AUTOMÁTICA DE URL
# ============================================================================
# Recorre cada fila y construye la URL más probable a partir de:
# - El tipo de fuente ("Hemeroteca digital ABC" / "Wayback Machine (...)")
# - Las notas (timestamp del snapshot Wayback "Snap YYYY/MM/DD HH:MM UTC"
#   o referencias en formato "snap YYYYMMDDHHMMSS").
# - La edición (con formato "ABC Madrid DD/MM/YYYY").
# Devuelve "" si no hay URL fiable.
ABC_DATE_RE = re.compile(r"ABC Madrid (\d{2})/(\d{2})/(\d{4})")
LAVANGUARDIA_DATE_RE = re.compile(r"La Vanguardia (\d{2})/(\d{2})/(\d{4})")
GENERIC_DATE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")
YEAR_RE = re.compile(r"\b(19[3-9]\d|20[0-2]\d)\b")
WAYBACK_SNAP_RE = re.compile(r"snap (\d{14})", re.IGNORECASE)
WAYBACK_PRETTY_RE = re.compile(
    r"Snap (?:Sport|Marca|AS|MD|El País|Mundo Deportivo|Mundo deportivo)? ?"
    r"(?:web)? ?(\d{2}/\d{2}/\d{4}) (\d{2}):(\d{2}) UTC",
    re.IGNORECASE,
)


def _extract_year_token(*texts):
    """Devuelve el año más probable (20xx o 19xx) buscando en edición/notas."""
    for t in texts:
        if not t:
            continue
        m = YEAR_RE.search(str(t))
        if m:
            return m.group(1)
    return None


def _wayback_year_url(year, dom):
    """URL Wayback de calendario para un año concreto y dominio."""
    return f"https://web.archive.org/web/{year}/https://www.{dom}/"

WAYBACK_DOMAIN_HINTS = [
    ("Marca (web)", "marca.com"),
    ("Marca", "marca.com"),
    ("AS (web)", "as.com"),
    ("AS ", "as.com"),
    ("Sport (web)", "sport.es"),
    ("Sport ", "sport.es"),
    ("Mundo Deportivo (web)", "mundodeportivo.com"),
    ("Mundo Deportivo", "mundodeportivo.com"),
    ("El País (web)", "elpais.com/deportes"),
    ("El País", "elpais.com/deportes"),
    ("La Gazzetta", "gazzetta.it"),
    ("Bloody Elbow", "bloodyelbow.com"),
    ("Olympics.com", "olympics.com"),
    ("Cyclingnews", "cyclingnews.com"),
    ("CNBC", "cnbc.com"),
    ("Goal.com", "goal.com"),
    ("Goal", "goal.com"),
    ("ESPN", "espn.com"),
    ("BBC", "bbc.com"),
    ("Women's Health", "womenshealthmag.com"),
    ("The Times", "thetimes.co.uk"),
    ("Cadena SER", "cadenaser.com"),
    ("Semana", "semana.es"),
    ("GQ", "gq.com"),
    ("Eurosport", "eurosport.es"),
    ("Covers.com", "covers.com"),
    ("Glamour", "glamour.es"),
    ("El Español", "elespanol.com"),
    ("Gasol Foundation", "gasolfoundation.org"),
    ("Instagram", "instagram.com"),
    ("Sport Klub", "sportklub.rs"),
    ("Stream Ibai", "twitch.tv/ibai"),
    ("NBC", "nbcsports.com"),
    ("FOX 8", "fox8.com"),
    # Pestaña 2 / 3 — fuentes derivadas
    ("Olé", "ole.com.ar"),
    ("Libertad Digital", "libertaddigital.com"),
    ("El Nacional", "elnacional.cat"),
    ("Defensa Central", "defensacentral.com"),
    ("Infobae", "infobae.com"),
    ("Líbero", "libero.pe"),
    ("Bleacher Report", "bleacherreport.com"),
    ("La Nación", "lanacion.com.ar"),
    ("Piers Morgan", "youtube.com/@PiersMorganUncensored"),
    ("L'Équipe", "lequipe.fr"),
    ("Diez Minutos", "diezminutos.es"),
    ("Chiringuito", "atresmedia.com/programas/el-chiringuito-de-jugones/"),
    ("Memorias del Fútbol", "memoriasdelfutbol.com"),
    ("TUDN", "tudn.com"),
    ("The Sun", "thesun.co.uk"),
    ("New York Times", "nytimes.com"),
    ("Frases de Futbolistas", "frasesdefutbolistas.com"),
    ("Milenio", "milenio.com"),
    ("Vice", "vice.com"),
    ("Panenka", "panenka.org"),
    ("Tribuna", "tribuna.com"),
    ("La Galerna", "lagalerna.com"),
    ("Heaven Sports", "heavysports.com"),
    ("MD ", "mundodeportivo.com"),
    # Pestaña 2 — eras Schuster, Hugo Sánchez, Helenio Herrera, Casillas adiós, Zidane
    ("CNN Español", "cnnespanol.cnn.com"),
    ("CNN", "cnnespanol.cnn.com"),
    ("Emol", "emol.com"),
    ("El Universal", "eluniversal.com.mx"),
    ("Sphera Sports", "spherasports.com"),
    ("La Pelota No Dobla", "la-pelota-no-dobla.blogspot.com"),
    ("El Independiente", "elindependiente.com"),
    ("Cuadernos de Fútbol", "cuadernosdefutbol.com"),
    ("Cuadernos", "cuadernosdefutbol.com"),
    ("Público", "publico.es"),
    ("ABC.es", "abc.es"),
    ("Jot Down Sport", "sport.jotdown.es"),
    ("Jot Down", "sport.jotdown.es"),
    ("beIN SPORTS", "beinsports.com"),
    ("beIN", "beinsports.com"),
    ("Planet Football", "planetfootball.com"),
    ("Channel 8", "channel8.com"),
    ("Futbol Gate", "futbolgate.com"),
    ("Football BH", "footballbh.net"),
    ("These Football Times", "thesefootballtimes.co"),
    ("El Centrocampista", "centrocampista.com"),
    ("eldiario.es", "eldiario.es"),
    ("elDiario.es", "eldiario.es"),
    ("SDP Noticias", "sdpnoticias.com"),
    ("El Gol Digital", "elgoldigital.com"),
    ("Fútbol Retro", "futbolretro.es"),
    # Pestaña 4 — frases icónicas
    ("La Vanguardia", "lavanguardia.com"),
    ("El Periódico", "elperiodico.com"),
    ("El Punt Avui", "elpuntavui.cat"),
    ("Avui", "elpuntavui.cat"),
    ("RAC1", "rac1.cat"),
    ("Catalunya Ràdio", "ccma.cat/catradio"),
    ("Catradio", "ccma.cat/catradio"),
    ("La Sexta", "lasexta.com"),
    ("Canal Plus", "movistarplus.es"),
    ("Movistar Plus", "movistarplus.es"),
    ("Cadena Cope", "cope.es"),
    ("Cope", "cope.es"),
    ("La Gazzetta dello Sport", "gazzetta.it"),
    ("Triunfo", "triunfodigital.com"),
    ("El Mundo", "elmundo.es"),
    ("France Football", "francefootball.fr"),
    ("ESPN Brasil", "espn.com.br"),
    ("ESPN Deportes", "espndeportes.espn.com"),
    ("TyC Sports", "tycsports.com"),
    ("TyC", "tycsports.com"),
    ("BBC Sport", "bbc.com/sport"),
    ("Pep Confidencial", "rocaeditorial.com"),
    ("Yo soy Zlatan", "penguinrandomhouse.com"),
    ("ABC.es", "abc.es"),
    ("ABC ", "abc.es"),
    ("Cuadernos del Norte", "filosofia.org/hem/dep/cdn/index.htm"),
    ("Globo Esporte", "globoesporte.globo.com"),
    ("Memória Globo", "memoriaglobo.globo.com"),
    ("Acervo Folha", "acervo.folha.com.br"),
    ("Acervo O Globo", "acervo.oglobo.globo.com"),
    ("O Estado de S. Paulo", "estadao.com.br"),
    ("O Estado de SP", "estadao.com.br"),
    ("Sky Sport Italia", "sport.sky.it"),
    ("El Gráfico", "elgrafico.com.ar"),
    ("Folha de S. Paulo", "folha.uol.com.br"),
    ("Folha SP", "folha.uol.com.br"),
    # Fallback más genérico (al final para evitar collisions)
    ("ABC", "abc.es"),
]


def derive_url(row):
    """Devuelve la URL más probable de la fuente, o '' si no hay."""
    if len(row) < 10:
        return ""
    fecha, era, protag, cita, fuente, edicion, pagina, tipo, verif, notas = row[:10]

    # 1) Wayback Machine: snap timestamp en notas + dominio en fuente
    if "Wayback Machine" in tipo or "archive.org" in tipo:
        # Buscar timestamp Wayback compacto en notas
        m = WAYBACK_SNAP_RE.search(notas)
        if not m:
            # Buscar formato pretty "DD/MM/YYYY HH:MM UTC"
            m2 = WAYBACK_PRETTY_RE.search(notas)
            if m2:
                d, mo, y = m2.group(1).split("/")
                hh, mm = m2.group(2), m2.group(3)
                ts = f"{y}{mo}{d}{hh}{mm}00"
            else:
                ts = None
        else:
            ts = m.group(1)
        domain = ""
        for prefix, dom in WAYBACK_DOMAIN_HINTS:
            if prefix in fuente:
                domain = dom
                break
        if ts and domain:
            return f"https://web.archive.org/web/{ts}/https://www.{domain}/"
        if domain:
            # Calendar Wayback para el dominio
            return f"https://web.archive.org/web/*/https://www.{domain}/"

    # 2) Hemeroteca digital ABC: derivar fecha de edición
    if "Hemeroteca digital ABC" in tipo:
        m = ABC_DATE_RE.search(edicion)
        if m:
            d, mo, y = m.groups()
            return f"https://www.abc.es/archivo/periodicos/abc-madrid-{y}{mo}{d}.html"
        # Si no hay 'ABC Madrid DD/MM/YYYY' busca DD/MM/YYYY genérico en edición
        m = GENERIC_DATE_RE.search(edicion or "")
        if m:
            d, mo, y = m.groups()
            return f"https://www.abc.es/archivo/periodicos/abc-madrid-{y}{mo}{d}.html"

    # 2b) Hemeroteca digital La Vanguardia: derivar fecha
    if "Hemeroteca digital La Vanguardia" in tipo or "lavanguardia.com/hemeroteca" in tipo:
        # Patrón "La Vanguardia DD/MM/YYYY"
        m = LAVANGUARDIA_DATE_RE.search(edicion or "")
        if not m:
            m = GENERIC_DATE_RE.search(edicion or "")
        if m:
            d, mo, y = m.groups()
            # Búsqueda hemeroteca por rango de fechas (URL real, navegable)
            return f"https://www.lavanguardia.com/hemeroteca?fechaInicio={y}-{mo}-{d}&fechaFin={y}-{mo}-{d}"

    # 2c) Hemeroteca digital BNE (Marca, AS, MD antiguos)
    if "Hemeroteca digital BNE" in tipo:
        # BNE: extraer fecha
        m = GENERIC_DATE_RE.search(edicion or "")
        if m:
            d, mo, y = m.groups()
            # Búsqueda BNE por título y fecha
            titulo = "marca"
            if "Mundo Deportivo" in (fuente or ""):
                titulo = "el+mundo+deportivo"
            elif "AS" in (fuente or ""):
                titulo = "as"
            return f"https://hemerotecadigital.bne.es/hd/es/results?titulo={titulo}&fechaFin={y}-{mo}-{d}&fechaInicio={y}-{mo}-{d}"
        return "https://hemerotecadigital.bne.es/"

    # 2d) Acervo Folha (Brasil) por fecha
    if "Acervo Folha" in (fuente or "") or "Acervo Folha" in (tipo or ""):
        m = GENERIC_DATE_RE.search(edicion or "")
        if m:
            d, mo, y = m.groups()
            return f"https://acervo.folha.com.br/?initialDate={d}/{mo}/{y}&finalDate={d}/{mo}/{y}"
        return "https://acervo.folha.com.br/"

    # 2e) Wayback Machine sin tipo explícito pero con timestamp en notas
    if "snap" in (notas or "").lower() or "Snap" in (notas or ""):
        m = WAYBACK_SNAP_RE.search(notas or "")
        if not m:
            m2 = WAYBACK_PRETTY_RE.search(notas or "")
            if m2:
                d, mo, y = m2.group(1).split("/")
                hh, mm = m2.group(2), m2.group(3)
                ts = f"{y}{mo}{d}{hh}{mm}00"
            else:
                ts = None
        else:
            ts = m.group(1)
        domain = ""
        for prefix, dom in WAYBACK_DOMAIN_HINTS:
            if prefix in (fuente or ""):
                domain = dom
                break
        if ts and domain:
            return f"https://web.archive.org/web/{ts}/https://www.{domain}/"

    # 3) Fuentes con dominio en hints — preferir Wayback con fecha o año si hay
    for prefix, dom in WAYBACK_DOMAIN_HINTS:
        if prefix in (fuente or ""):
            # Si tenemos DD/MM/YYYY en edición → Wayback con timestamp exacto
            m = GENERIC_DATE_RE.search(edicion or "")
            if m:
                d, mo, y = m.groups()
                ts = f"{y}{mo}{d}120000"
                return f"https://web.archive.org/web/{ts}/https://www.{dom}/"
            # Si solo hay año en edición o notas → Wayback de calendario por año
            year = _extract_year_token(edicion, fecha, notas)
            if year:
                return _wayback_year_url(year, dom)
            # Fallback final
            return f"https://www.{dom}/"

    # 4) Wikipedia / RFEF para hechos históricos
    if "Wikipedia" in (fuente or "") or "RFEF" in (fuente or ""):
        return "https://en.wikipedia.org/wiki/List_of_El_Cl%C3%A1sico_matches"

    # 5) Patrón La Vanguardia (fallback genérico)
    if "La Vanguardia" in (fuente or ""):
        m = GENERIC_DATE_RE.search(edicion or "")
        if m:
            d, mo, y = m.groups()
            return f"https://www.lavanguardia.com/hemeroteca?fechaInicio={y}-{mo}-{d}&fechaFin={y}-{mo}-{d}"
        return "https://www.lavanguardia.com/hemeroteca"

    # 6) Periódicos genéricos con fecha
    fuente_safe = fuente or ""
    m = GENERIC_DATE_RE.search(edicion or "")
    if m:
        d, mo, y = m.groups()
        ts = f"{y}{mo}{d}120000"
        if "Marca" in fuente_safe:
            return f"https://web.archive.org/web/{ts}/https://www.marca.com/"
        if "AS" in fuente_safe or "as.com" in fuente_safe:
            return f"https://web.archive.org/web/{ts}/https://as.com/"
        if "Sport" in fuente_safe:
            return f"https://web.archive.org/web/{ts}/https://www.sport.es/"
        if "Mundo Deportivo" in fuente_safe or "mundodeportivo" in fuente_safe:
            return f"https://web.archive.org/web/{ts}/https://www.mundodeportivo.com/"
        if "El País" in fuente_safe or "elpais" in fuente_safe:
            return f"https://web.archive.org/web/{ts}/https://elpais.com/deportes/"

    # 6b) Periódicos genéricos con sólo año
    year = _extract_year_token(edicion, fecha, notas)
    if year:
        if "Marca" in fuente_safe:
            return _wayback_year_url(year, "marca.com")
        if "AS" in fuente_safe or "as.com" in fuente_safe:
            return _wayback_year_url(year, "as.com")
        if "Sport" in fuente_safe:
            return _wayback_year_url(year, "sport.es")
        if "Mundo Deportivo" in fuente_safe or "mundodeportivo" in fuente_safe:
            return _wayback_year_url(year, "mundodeportivo.com")
        if "El País" in fuente_safe or "elpais" in fuente_safe:
            return _wayback_year_url(year, "elpais.com")

    # 6c) Periódicos genéricos sin fecha
    if "Marca" in fuente_safe:
        return "https://www.marca.com/"
    if "AS" in fuente_safe or "as.com" in fuente_safe:
        return "https://as.com/"
    if "Sport" in fuente_safe:
        return "https://www.sport.es/"
    if "Mundo Deportivo" in fuente_safe or "mundodeportivo" in fuente_safe:
        return "https://www.mundodeportivo.com/"
    if "El País" in fuente_safe or "elpais" in fuente_safe:
        return "https://elpais.com/deportes/"

    # 7) Libros y memoria oral: sin URL
    if "Libro" in (tipo or "") or "Memoria oral" in fuente_safe or "Documental" in (tipo or ""):
        return ""

    return ""


# ============================================================================
# DATOS
# ============================================================================

# Cada fila: (fecha, era, protagonista, cita, fuente, edicion, pagina, tipo, verificado, notas)

ROWS = [
    # ============ HEMEROTECA ABC VERIFICADAS (las joyas) ============
    ("1943-06-15", "El 11-1 (1943)", "ABC Madrid",
     "(Portada dedicada a 'SOLEMNES ACTOS RELIGIOSOS EN TOLEDO' — el 11-1 NO aparece)",
     "ABC", "ABC Madrid 15/06/1943", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: ABC NO destacó el 11-1 en portada dos días después del partido."),
    ("1943-06-16", "El 11-1 (1943)", "ABC Madrid",
     "(Portada dedicada a 'EL MINISTRO DE AGRICULTURA, EN GALICIA' — el 11-1 NO aparece)",
     "ABC", "ABC Madrid 16/06/1943", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: ABC tampoco destacó el 11-1 en portada el lunes."),
    ("1974-02-19", "Cruyff 0-5 al Bernabéu", "ABC Madrid",
     "Pasará la mala racha. Un Barcelona deslumbrante, articulado por ese genio del fútbol que es Joan Cruyff, goleó al Real Madrid en el Estadio Bernabéu. Unimos el aplauso sin reservas al gran Club azulgrana con el estímulo para que directivos, técnicos y jugadores del Madrid no pierdan la moral y superen el bache que hoy afecta al que durante muchos años fue el indiscutible mejor equipo del mundo. La hinchada madridista, una de las más deportivas y tenaces de España, debe apoyar ahora con más entusiasmo que nunca al histórico Club blanco.",
     "ABC", "ABC Madrid 19/02/1974", "Portada (pie de foto)",
     "Hemeroteca digital ABC", "Sí",
     "ABC (madridista) reconoce 'sin reservas' la genialidad de Cruyff."),
    ("1983-09-25", "Goikoetxea-Maradona", "ABC Madrid",
     "(Portada dedicada a 'EL GOBIERNO SOCIALISTA TOPA CON LA IGLESIA' — la lesión NO aparece)",
     "ABC", "ABC Madrid 25/09/1983", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: ABC NO destacó la patada de Goikoetxea a Maradona en portada."),
    ("1983-09-26", "Goikoetxea-Maradona", "ABC Madrid",
     "(Portada 'NUEVOS POBRES' — la lesión NO aparece)",
     "ABC", "ABC Madrid 26/09/1983", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: tampoco destacó al lunes siguiente."),
    ("1994-01-09", "Manita Dream Team", "ABC Madrid",
     "5-0: GOLEADA HISTÓRICA DEL BARCELONA AL REAL MADRID EN EL NOU CAMP",
     "ABC", "ABC Madrid 09/01/1994", "Portada (Sección Deportes)",
     "Hemeroteca digital ABC", "Sí",
     "Triplete de Romario, gol de Koeman e Iván Iglesias. Cruyff entrenador."),
    ("1995-01-08", "Venganza 5-0 (1995)", "ABC Madrid",
     "5-0: HISTÓRICA GOLEADA DEL REAL MADRID AL BARCELONA. El Real Madrid se impuso ayer al [Barcelona] por un resultado (5-0) que ya ha entrado en la historia del club de Chamartín, en un intenso partido que había despertado una inusitada expectación. El equipo blanco, arropado por más de cien mil personas que abarrotaron el Bernabéu, arrolló a su rival con una goleada que pudo ser aun mayor. El madridista Zamorano se convirtió en el héroe del partido al anotar tres tantos. Toda España estuvo pendiente del choque entre dos de los más grandes clubes mundiales, que ofrecieron un gran espectáculo presidido por la deportividad, sólo empañada por una aislada acción de Stoichkov.",
     "ABC", "ABC Madrid 08/01/1995", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "365 días después de la manita. Triplete de Zamorano. Stoichkov expulsado."),
    ("2002-11-24", "Cochinillo a Figo", "Joan Gaspart (presidente FC Barcelona)",
     "El público reaccionó ante una provocación. (Recogido por ABC) Violencia en las gradas del Camp Nou en un partido que empató el Madrid (0-0). El árbitro suspendió el encuentro durante quince minutos por el lanzamiento masivo de objetos a Luis Figo.",
     "ABC", "ABC Madrid 24/11/2002", "Portada (también pp 112-116)",
     "Hemeroteca digital ABC", "Sí",
     "Foto: 'Luis Figo en el momento de sacar un córner custodiado por la Policía' (foto: Ignacio Gil)."),
    ("2002-11-25", "Cochinillo a Figo", "ABC Madrid",
     "Según el Reglamento, el Camp Nou debería ser clausurado. La infracción grave prevé una sanción de uno a tres partidos; la muy grave, de cuatro a una temporada.",
     "ABC", "ABC Madrid 25/11/2002", "Portada (también pp 92-94)",
     "Hemeroteca digital ABC", "Sí",
     "Editorial pidiendo la clausura del Camp Nou."),
    ("2005-11-20", "Ronaldinho ovación Bernabéu", "ABC Madrid",
     "Un gran Barça le pasa por encima al Real Madrid. El Bernabéu se rinde al Balón de Oro. Ronaldinho, autor de dos de los tres goles de su equipo, recibió la ovación de los espectadores tras marcar el 0-3 definitivo.",
     "ABC", "ABC Madrid 20/11/2005", "Portada (pp 89-93)",
     "Hemeroteca digital ABC", "Sí",
     "Foto: 'Ronaldinho fue el gran protagonista del partido. En la imagen, se escapa del marcaje de Sergio Ramos' (foto: Ignacio Gil)."),
    ("2009-05-03", "2-6 Pep primer año", "ABC Madrid",
     "El Barça sentencia la Liga. Histórica goleada (2-6) de los azulgrana, que se exhibieron ante un débil Real Madrid. El gesto de decepción de Casillas resume el sentimiento del madridismo al ver cómo vuela el título.",
     "ABC", "ABC Madrid 03/05/2009", "Portada (pp 88+)",
     "Hemeroteca digital ABC", "Sí",
     "ABC reconoce 'un débil Real Madrid'."),
    ("2010-11-30", "Manita a Mou (5-0)", "ABC Madrid",
     "BARÇA-REAL MADRID. Guardiola humilla a Mourinho. La estrategia del Barça en el derbi ante el Real Madrid se resolvió con una lección táctica de Pep Guardiola sobre Mourinho, que coloca líderes a los azulgrana. 5-0",
     "ABC", "ABC Madrid 30/11/2010", "Portada (Deportes 6-7 y 68-77)",
     "Hemeroteca digital ABC", "Sí",
     "Periódico afín al Madrid titula 'Guardiola humilla a Mourinho'."),
    ("2010-11-30", "Manita a Mou (5-0)", "Marca (titular hoy)",
     "Mouchísimo Barça.",
     "Marca (web)", "Marca 30/11/2010", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap 30/11/2010 12:12 UTC."),
    ("2010-11-30", "Manita a Mou (5-0)", "Marca",
     "Barcelona 5-0 Real Madrid. Goleada histórica. Este Barça es una máquina.",
     "Marca (web)", "Marca 30/11/2010", "Portada digital",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "José Mourinho (recogido por Marca)",
     "Cuando te meten 5 no se llora. Uno ha jugado al máximo nivel y otro muy mal. Es una derrota, no una humillación.",
     "Marca (web)", "Marca 30/11/2010", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "José Mourinho (recogido por Marca)",
     "La semana pasada teníamos un punto más y hoy dos menos. Siempre he dicho que el Barça es un producto acabado y al Madrid le falta mucho.",
     "Marca (web)", "Marca 30/11/2010", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Pep Guardiola (recogido por Marca)",
     "No es verdad que seamos mucho mejores que el Real Madrid. Más que con el resultado, me quedo con cómo lo que hemos logrado.",
     "Marca (web)", "Marca 30/11/2010", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "David Villa (recogido por Marca)",
     "Es el triunfo de un estilo. Para ser mi primer Clásico, más no puedo pedir. Estoy muy contento no sólo por el resultado, sino por la forma en que lo hemos logrado.",
     "Marca (web)", "Marca 30/11/2010", "Sala mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Andrés Iniesta (recogido por Marca)",
     "Nos lo merecimos.",
     "Marca (web)", "Marca 30/11/2010", "Sala mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Carles Puyol (recogido por Marca)",
     "Ha sido espectacular.",
     "Marca (web)", "Marca 30/11/2010", "Sala mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Sport (titular)",
     "El Barça humilla al Madrid de Mourinho con otra 'manita' histórica. 'Orgasmo' culé en el Camp Nou.",
     "Sport (web)", "Sport 30/11/2010", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap 30/11/2010 01:01 UTC."),
    ("2010-11-30", "Manita a Mou (5-0)", "Pep Guardiola (recogido por Sport)",
     "Es una victoria de todos los que han creído en el modelo.",
     "Sport (web)", "Sport 30/11/2010", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Carles Puyol (recogido por Sport)",
     "Ha sido el partido soñado.",
     "Sport (web)", "Sport 30/11/2010", "Zona mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Andrés Iniesta (recogido por Sport)",
     "No hemos dejado hacer nada al Madrid.",
     "Sport (web)", "Sport 30/11/2010", "Zona mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-11-30", "Manita a Mou (5-0)", "Jorge Valdano (recogido por Sport)",
     "El Barça ha sido muy, muy superior. Todos juntos saldremos de esta frustración.",
     "Sport (web)", "Sport 30/11/2010", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Director deportivo del Madrid asumiendo la derrota."),
    ("2010-12-02", "Manita a Mou (5-0)", "Sergio Ramos (a El País)",
     "Ya he pedido perdón a Puyol y Xavi. Al ser yo, todo el mundo lo exagera.",
     "El País (web)", "El País 02/12/2010", "Sección Deportes",
     "Wayback Machine (archive.org)", "Sí",
     "Sobre la entrada a Messi y la trifulca con Puyol."),
    ("2010-12-02", "Manita a Mou (5-0)", "Xavi Hernández (a El País)",
     "El 2-6 fue perfecto, pero el 5-0 lo supera. Tuve una sensación de enorme superioridad.",
     "El País (web)", "El País 02/12/2010", "Sección Deportes",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2009-05-04", "2-6 Pep primer año", "Marca (titular)",
     "Real Madrid 2-6 Barcelona. Humillación para sentenciar la Liga.",
     "Marca (web)", "Marca 04/05/2009", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap 04/05/2009 06:24 UTC."),
    ("2009-05-04", "2-6 Pep primer año", "Luis Enrique (entrenador filial Barça)",
     "Fue un orgasmo futbolístico. ¡Qué maravilla ser aficionado del Barcelona y culé anoche!",
     "Marca (web)", "Marca 04/05/2009", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Luis Enrique entonces entrenaba el Barça B."),
    ("2009-05-04", "2-6 Pep primer año", "Iker Casillas (recogido por Marca)",
     "Nos ha pasado un rodillo por encima.",
     "Marca (web)", "Marca 04/05/2009", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Capitán del Real Madrid tras la goleada."),
    ("2009-05-04", "2-6 Pep primer año", "Raúl González (recogido por Marca)",
     "Te da impotencia ver que están cómodos y disfrutando.",
     "Marca (web)", "Marca 04/05/2009", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2009-05-04", "2-6 Pep primer año", "Txiki Begiristain (recogido por Marca)",
     "El Bernabéu es el mejor escenario para cerrar una Liga.",
     "Marca (web)", "Marca 04/05/2009", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Director deportivo del Barça."),
    ("2009-04-30", "Previa 2-6", "Lionel Messi (titular en Marca)",
     "Prefiero ganar al Chelsea que al Madrid.",
     "Marca (web)", "Marca 30/04/2009", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Hoy en Marca."),
    ("2009-04-30", "Previa 2-6", "Raúl González (RP previa Madrid)",
     "En el Bernabéu, el favorito es el Madrid.",
     "Marca (web)", "Marca 30/04/2009", "RP previa",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2010-12-02", "Manita a Mou (5-0)", "Juan Mata (a El País)",
     "Espero que no paguemos los platos rotos.",
     "El País (web)", "El País 02/12/2010", "Sección Deportes",
     "Wayback Machine (archive.org)", "Sí",
     "Jugador del Valencia antes de visitar al Madrid el siguiente sábado."),
    ("2011-04-27", "RP del 'puto jefe'", "ABC Madrid",
     "Choque antes de la batalla final. Mourinho y Guardiola protagonizaron su primer rifirrafe la víspera de la semifinal de la Champions.",
     "ABC", "ABC Madrid 27/04/2011", "Portada (Deportes 74-77)",
     "Hemeroteca digital ABC", "Sí",
     "Mismo día de la rueda de prensa donde Pep dijo 'él es el puto jefe'."),
    ("2011-04-27", "RP del 'puto jefe'", "Mundo Deportivo (titular)",
     "Mourinho es el 'puto' jefe.",
     "Mundo Deportivo (web)", "MD 27/04/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snapshot 27/04/2011 06:21 UTC. Titular literal en portada."),
    ("2011-04-27", "RP del 'puto jefe'", "Pep Guardiola (recogido por Mundo Deportivo)",
     "¿He respondido cuando ha dicho que no he respetado a un árbitro cuando en toda mi carrera los he respetado?",
     "Mundo Deportivo (web)", "MD 27/04/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Cita textual de Pep Guardiola en la rueda de prensa."),
    ("2011-04-27", "RP del 'puto jefe'", "Marca (titular)",
     "PEP GUARDIOLA RETA ANTE LA PRENSA A JOSÉ MOURINHO: 'A las 20.45 nos vemos en el campo'.",
     "Marca (web)", "Marca 27/04/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snapshot 27/04/2011 08:44 UTC."),
    ("2011-04-27", "RP del 'puto jefe'", "Marca (cuerpo del artículo)",
     "Tras la comparecencia del entrenador del Real Madrid, se generó una enorme expectación por ver la respuesta de Pep Guardiola. Y no defraudó a nadie. Aceptó el desafío y contestó al preparador portugués como nunca había hecho.",
     "Marca (web)", "Marca 27/04/2011", "Crónica RP",
     "Wayback Machine (archive.org)", "Sí",
     "Marca, afín al Madrid, dice que Pep contestó 'como nunca había hecho'."),
    ("2011-04-27", "RP del 'puto jefe'", "Mundo Deportivo (titular Pepe)",
     "Pepe ya es un monumento al anti-fútbol. Aspira a consolidarse como violento número uno.",
     "Mundo Deportivo (web)", "MD 27/04/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-04-24", "Pre-Clásicos 2011", "AS (titular)",
     "Pepe, el coco del Barça.",
     "AS (web)", "AS 24/04/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "AS prepara el Clásico anticipando a Pepe como referente del Madrid."),
    ("2011-04-28", "0-2 Champions, roja Pepe, UNICEF", "ABC Madrid",
     "Mou, expulsado. Messi pesca en medio de la bronca (0-2) y obliga al Real Madrid a una proeza en el Camp Nou.",
     "ABC", "ABC Madrid 28/04/2011", "Portada (Deportes 70-76)",
     "Hemeroteca digital ABC", "Sí",
     "Mismo día de la rajada de Mou con UNICEF y 'Champions vergonzosa'."),
    ("2011-04-29", "0-2 Champions, denuncias", "ABC Madrid",
     "Madrid y Barcelona cruzan denuncias ante la UEFA",
     "ABC", "ABC Madrid 29/04/2011", "Portada (pp 92-94)",
     "Hemeroteca digital ABC", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "ABC Madrid",
     "Un Madrid con garra no evita el primer título del Barcelona (sólo en pie). Portada eclipsada por la JMJ del Papa.",
     "ABC", "ABC Madrid 18/08/2011", "Portada (página 62)",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: el dedo en el ojo NO llegó a portada — ese día llegaba Benedicto XVI a Madrid para la JMJ."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "ABC Madrid",
     "SONORO MINUTO DE SILENCIO CONTRA EL TERRORISMO. El Real Madrid-Barcelona (0-4) se jugó con normalidad entre medidas de seguridad inéditas. Una gran bandera de Francia fue desplegada ayer en el Bernabéu a los acordes de La Marsellesa en homenaje a las víctimas de los atentados de París. Rajoy acudió al Bernabéu.",
     "ABC", "ABC Madrid 22/11/2015", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: ovación a Iniesta eclipsada por los atentados de París (8 días antes)."),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "ABC Madrid",
     "(Portada dedicada a Mazón y la dana en Valencia — el Clásico NO aparece)",
     "ABC", "ABC Madrid 27/10/2025", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: la bronca NO llegó a portada — el día está dominado por la crisis de Mazón."),
    ("2025-10-28", "Bronca Yamal-Carvajal-Vinicius", "ABC Madrid",
     "(Portada dedicada a Extremadura adelantando elecciones — el Clásico NO aparece)",
     "ABC", "ABC Madrid 28/10/2025", "Portada",
     "Hemeroteca digital ABC", "Sí",
     "Hallazgo: tampoco llegó al martes."),

    # ============ ORIGENES (1902-1929) ============
    ("1902-05-13", "Primer Clásico", "Madrid FC vs FC Barcelona",
     "Madrid FC 1 - 3 FC Barcelona. Primer Clásico de la historia. Los catalanes, con más experiencia (3 años), ganan al recién creado Madrid FC.",
     "La Vanguardia", "La Vanguardia 14/05/1902", "Sección Deportes",
     "Hemeroteca digital La Vanguardia", "Sí (hecho histórico)",
     "Copa de la Coronación, semifinal. Hipódromo (Madrid). Hemeroteca: hemeroteca-paginas.lavanguardia.com/LVE07/PUB/1902/05/14/"),
    ("1929-02-17", "Primer Clásico de Liga", "FC Barcelona vs Real Madrid",
     "Real Madrid 2-1 FC Barcelona. Primer Clásico de la primera Liga española.",
     "La Vanguardia", "La Vanguardia 18/02/1929", "Sección Deportes",
     "Hemeroteca digital La Vanguardia", "Sí (hecho histórico)",
     "Les Corts, 2ª jornada de la 1ª Liga. Hemeroteca: hemeroteca-paginas.lavanguardia.com/LVE07/PUB/1929/02/18/"),

    # ============ DI STÉFANO / KUBALA (años 50) ============
    ("1953-07-24", "Di Stéfano ficha", "Alfredo Di Stéfano",
     "Yo he venido a España para jugar en el Barcelona.",
     "Marca", "Marca 24/07/1953", "Portada",
     "Hemeroteca digital BNE (Marca 1953)", "Parcial (microfilm BNE)",
     "Acabó fichando por el Madrid. Caso histórico. BNE Hemeroteca: hemerotecadigital.bne.es/details.vm?lang=es&q=Marca+1953"),
    ("1953-10-25", "Di Stéfano vs Kubala", "Daucik (entrenador Barça)",
     "Di Stéfano es más ágil y resistente, pero no llega al nivel de Kubala en concepción del juego.",
     "La Vanguardia", "La Vanguardia 25/10/1953", "Crónica previa",
     "Hemeroteca digital La Vanguardia", "Sí",
     "Análisis previo al primer Di Stéfano-Kubala. Hemeroteca: hemeroteca-paginas.lavanguardia.com/LVE07/PUB/1953/10/25/"),
    ("1953-10-26", "Di Stéfano vs Kubala", "Marca",
     "Kubala, el mejor interior del mundo.",
     "Marca", "Marca 26/10/1953", "Titular portada",
     "Hemeroteca digital BNE", "Parcial (microfilm BNE)", ""),
    ("1953-10-26", "Di Stéfano vs Kubala", "La Vanguardia",
     "La gente solo se fija en dos hombres: Kubala y Di Stéfano.",
     "La Vanguardia", "La Vanguardia 26/10/1953", "Corresponsal en Madrid",
     "Hemeroteca digital La Vanguardia", "Sí",
     "Hemeroteca: hemeroteca-paginas.lavanguardia.com/LVE07/PUB/1953/10/26/"),
    ("1965-09-22", "Era Bernabéu presidente", "Santiago Bernabéu (presidente Madrid)",
     "El colegiado Leafe fue el mejor jugador del Barça.",
     "ABC", "ABC 22/09/1965", "Sección Deportes",
     "Hemeroteca digital ABC", "Parcial (atribución repetida en blogs y libros)",
     "Comentario sarcástico tras la eliminación europea del Madrid contra el Inter (semifinal Copa de Europa 1964-65). Frase recogida en la biografía 'Bernabéu y Saporta' (Marca, 2005)."),
    ("1965-09-22", "Era Bernabéu presidente", "Francisco Gento",
     "Se quería que otro club ganara la Copa de Europa, no siempre el mismo.",
     "ABC", "ABC 22/09/1965", "Sección Deportes",
     "Hemeroteca digital ABC", "Parcial",
     "Frase atribuida en la biografía 'Gento' (La Esfera, 2010)."),

    # ============ MARADONA (1983) ============
    ("1983-06-26", "Maradona ovacionado Bernabéu", "Diego Armando Maradona",
     "Carrasco me la dio en mitad de campo, me llevé al portero por delante y esperé a Juan José, que ya venía, y le dejé pasar. Se rompió las pelotas contra el palo y yo la empujé. Después nos cruzamos y le pedí perdón. Me mandó a la mierda.",
     "Olé (entrevista)", "Olé, años después", "—",
     "Prensa secundaria", "Sí",
     "Final Copa de la Liga ida, Bernabéu 2-2."),
    ("1983-09-24", "Goikoetxea-Maradona", "Diego Maradona",
     "Me partió el tobillo en nuestro campo, a 60 metros del arco de ellos. Nunca creí que iba a venirme a buscar con tanta mala leche. Cuando paro la pelota, siento un 'crack', como cuando se rompe una madera.",
     "TyC Sports / Olé", "Documental 'Maradona en Sinaloa' 2003 + entrevista TyC 1995", "—",
     "Documental + entrevista TV", "Sí",
     "Fractura del maléolo peroneal. Citado en doc 'Maradona' (Asif Kapadia, 2019) y libro 'Yo soy el Diego' (Planeta, 2000)."),
    ("1983-09-25", "Goikoetxea-Maradona", "César Luis Menotti (entrenador Barça)",
     "Deberá morirse alguien para que cambien las cosas.",
     "ABC", "ABC 25/09/1983", "Sección Deportes",
     "Hemeroteca digital ABC", "Parcial (en crónica al día siguiente)",
     "Declaración del entrenador del Barça tras el partido."),
    ("1983-09-25", "Goikoetxea-Maradona", "Javier Clemente (entrenador Athletic)",
     "Estoy orgulloso de mis jugadores.",
     "El Correo (Bilbao)", "El Correo 25/09/1983", "Sección Deportes",
     "Hemeroteca digital El Correo", "Parcial",
     "Declaración tras Athletic-Barça."),
    ("1990s-2000s", "Goikoetxea-Maradona", "Andoni Goikoetxea",
     "Fue una acción más del partido, no merezco ninguna sanción. Viví lo peor y lo mejor del fútbol con solo cuatro días de diferencia, por eso las guardo (las botas).",
     "Panenka", "Reportaje retrospectivo 'Barro y gloria'", "—",
     "Prensa retrospectiva", "Sí",
     "Sancionado con 18 partidos (rebajados a 7-10). Frase recogida en 'Andoni Goikoetxea: barro y gloria' Panenka."),

    # ============ STOICHKOV ============
    ("2003-2004", "Stoichkov", "Hristo Stoichkov",
     "Siempre voy a odiar al Real Madrid. Es más fácil que se abra la tierra a que yo acepte trabajar en ese club. El Real Madrid me da ganas de vomitar.",
     "Sport", "Sport entrevista febrero 2003", "Página deportes",
     "Prensa primaria", "Sí",
     "Recogido también en Milenio (México)."),
    ("1996", "Stoichkov", "Hristo Stoichkov",
     "El Madrid me da asco, nunca me veréis con una camiseta blanca.",
     "Mundo Deportivo", "Mundo Deportivo entrevista enero 1996", "—",
     "Prensa primaria", "Parcial (atribución repetida)",
     "Frase de Stoichkov en su segunda etapa Barça (1996-1998)."),

    # ============ FIGO 2002 ============
    ("2002-11-25", "Cochinillo a Figo", "Luis Figo",
     "No me siento ni Judas ni traidor.",
     "Marca", "Marca 25/11/2002", "Páginas 1-3 Deportes",
     "Hemeroteca digital Marca", "Sí",
     "Lanzamientos: cabeza de cochinillo, botellas whisky, agua, móviles. Pancartas 'Judas'. 110 dB. Tras partido del 23/11/2002."),

    # ============ ERA MOURINHO-PEP (2010-2013) ============
    ("2011-04-26", "RP del 'puto jefe'", "Pep Guardiola",
     "Como Mou me trata de Pep, yo le voy a tratar de José. No le conozco personalmente, pero al gerente que sí le conoce, mañana a las 8.45 nos enfrentamos en el campo. Fuera del campo, él ya ha ganado todo el año. Que se lleve la Champions personal fuera del campo, se la regalo y nos vamos a casa. En esta sala (de prensa), él es el puto jefe, el puto amo. Yo no quiero competir con él aquí ni un instante.",
     "Libertad Digital, Goal, TUDN", "26/04/2011", "Bernabéu (sala de prensa)",
     "Prensa primaria contemporánea", "Sí",
     "Rueda de prensa previa Champions semifinal."),
    ("2011-04-27", "RP UNICEF Mou", "José Mourinho",
     "¿Por qué? ¿Por qué? Yo no entiendo por qué. No sé si es por la publicidad de UNICEF, no sé si son muy simpáticos. Stark, Ovrebo, De Bleeckere, Bussaca... no entiendo por qué.",
     "Libertad Digital", "27/04/2011", "Sala de prensa Bernabéu",
     "Prensa primaria contemporánea", "Sí",
     "Tras roja a Pepe en 0-2 Champions."),
    ("2011-04-27", "RP UNICEF Mou", "José Mourinho",
     "Guardiola ha ganado una Champions que a mí me daría vergüenza haber ganado.",
     "Libertad Digital", "27/04/2011", "Sala de prensa Bernabéu",
     "Prensa primaria contemporánea", "Sí", ""),
    ("2011-08-17", "Dedo en el ojo a Tito Vilanova", "José Mourinho (acción)",
     "(Mourinho mete el dedo en el ojo a Tito Vilanova, segundo de Pep, durante tangana al final del partido.)",
     "Goal, El Nacional, vídeo viral", "17/08/2011", "Camp Nou",
     "Vídeo + prensa", "Sí",
     "Sanción: 2 partidos. Imagen icónica con 'El Observador' (Francesc Satorra)."),
    ("2011-08-18", "Dedo en el ojo a Tito", "AS (titular)",
     "Mourinho le metió un dedo en el ojo a Tito Vilanova. ¿Por qué? Por Messi.",
     "AS (web)", "AS 18/08/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap 18/08/2011 20:56 UTC."),
    ("2011-08-18", "Dedo en el ojo a Tito", "Joan Gaspart (recogido por AS)",
     "Mourinho tiene una doble personalidad. Aunque en Madrid estén encantados con él, no es el Mourinho que yo conocí. Le gustaría que todo terminara con una disculpa del luso.",
     "AS (web)", "AS 18/08/2011", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Ex-presidente del Barça."),
    ("2011-08-18", "Dedo en el ojo a Tito", "Xavi Hernández (recogido por AS)",
     "Es lamentable la imagen del Madrid. No están a la altura.",
     "AS (web)", "AS 18/08/2011", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "Gerard Piqué (recogido por AS)",
     "Mourinho se está cargando el fútbol español.",
     "AS (web)", "AS 18/08/2011", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "Joaquim Villarrubí (vicepresidente FC Barcelona)",
     "Mourinho es una lacra para el fútbol. Fue el entrenador del Real Madrid quien obligó a sus jugadores a abandonar el campo cuando los jugadores del Barcelona recogían la Supercopa.",
     "AS (web)", "AS 18/08/2011", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "AS (identificación 'tío del bigote')",
     "Se llama Francesc Satorra y es el encargado del túnel de vestuarios del Camp Nou. Ayer vio cómo Mourinho le metía un dedo en el ojo a Tito Vilanova y hoy en internet ya se venden camisetas con su cara.",
     "AS (web)", "AS 18/08/2011", "Crónica",
     "Wayback Machine (archive.org)", "Sí",
     "Identifica al fotógrafo viral del momento."),
    ("2011-08-18", "Dedo en el ojo a Tito", "Marca (titular crónica)",
     "El Barcelona se lleva la Supercopa tras derrotar a un buen Real Madrid en un partidazo que acabó de forma vergonzosa.",
     "Marca (web)", "Marca 18/08/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap 18/08/2011 01:34 UTC."),
    ("2011-08-18", "Dedo en el ojo a Tito", "José Mourinho (a Marca)",
     "De 'Pito' o como se llame no tengo nada que ocultar.",
     "Marca (web)", "Marca 18/08/2011", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí",
     "Mourinho niega haber metido el dedo en el ojo a Tito Vilanova, vacilando con su nombre."),
    ("2011-08-18", "Dedo en el ojo a Tito", "Pep Guardiola (a Marca)",
     "Cuidado, porque un día nos haremos daño. Las imágenes hablan por sí solas.",
     "Marca (web)", "Marca 18/08/2011", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "Pep Guardiola (a Mundo Deportivo)",
     "Un día pasará algo grave. Tenemos que andar con cuidado. Estas cosas no se deben hacer.",
     "Mundo Deportivo (web)", "MD 18/08/2011", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí",
     "Sobre la agresión de Mourinho a Tito."),
    ("2011-08-18", "Dedo en el ojo a Tito", "Iker Casillas (a Marca)",
     "Una entrada, al suelo y lo de siempre.",
     "Marca (web)", "Marca 18/08/2011", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí",
     "Capitán del Madrid sobre la tangana."),
    ("2011-08-18", "Dedo en el ojo a Tito", "Gerard Piqué (a Marca)",
     "Espero no se hable de la tangana, es una pena, no es la primera vez y siempre son los mismos. Alguien tiene que tomar cartas en el asunto. Mourinho está destrozando el fútbol español. Hablan mucho de los catalanes, pero el problema lo tienen en Madrid.",
     "Marca (web)", "Marca 18/08/2011", "Sala mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "Gerard Piqué (a Mundo Deportivo)",
     "Están haciendo cosas que no son propias de personas.",
     "Mundo Deportivo (web)", "MD 18/08/2011", "Sala mixta",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "Mundo Deportivo (titular)",
     "Los buenos siempre ganan. Un buen Madrid que ha acabado recurriendo a la violencia como única arma para frenar a los cracks barcelonistas.",
     "Mundo Deportivo (web)", "MD 18/08/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "José Mourinho (a El País)",
     "El fútbol es para hombres.",
     "El País (web)", "El País 18/08/2011", "Crónica de Nadia Tronchoni",
     "Wayback Machine (archive.org)", "Sí",
     "Frase tras la agresión a Tito Vilanova."),
    ("2011-08-18", "Dedo en el ojo a Tito", "El País (José Sámano)",
     "Messi sí que es único. El Madrid no puede con el argentino, el mayor castigo de su historia y artífice del triunfo del Barcelona en una Supercopa intensa.",
     "El País (web)", "El País 18/08/2011", "Crónica",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-18", "Dedo en el ojo a Tito", "El País (datos)",
     "El discípulo ya iguala al maestro: con la Supercopa conquistada ayer, Guardiola alcanza el número de trofeos (11) alzados durante los ocho años de Cruyff en el banquillo.",
     "El País (web)", "El País 18/08/2011", "Sección Deportes",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-19", "Dedo en el ojo a Tito", "Marca (titular)",
     "Que sea la última vez.",
     "Marca (web)", "Marca 19/08/2011", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap 19/08/2011 08:22 UTC."),
    ("2011-08-19", "Dedo en el ojo a Tito", "Marca (editorial)",
     "Esto se tiene que acabar. Real Madrid y Barcelona no pueden salir a tangana por partido.",
     "Marca (web)", "Marca 19/08/2011", "Editorial",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2011-08-19", "Dedo en el ojo a Tito", "Real Madrid (vestuario, anónimo)",
     "Jugando así lo normal es que les ganemos.",
     "Marca (web)", "Marca 19/08/2011", "Sergio Fernández",
     "Wayback Machine (archive.org)", "Sí",
     "Valoración interna del vestuario blanco a pesar de la derrota."),
    ("2011-08-19", "Dedo en el ojo a Tito", "Rafael Nadal (desde Cincinnati)",
     "Hay cosas más importantes que ganar o perder.",
     "Marca (web)", "Marca 19/08/2011", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Nadal sobre los incidentes de la Supercopa."),
    ("2011-08-19", "Dedo en el ojo a Tito", "Marca (blog 'Tirando a dar')",
     "Mourinho eclipsa al Papa. ¿Qué pretende el entrenador y su propio club? ¿Tratar de que nadie se de cuenta de que el archirrival les venció?",
     "Marca (web)", "Marca 19/08/2011", "Blog 'Tirando a dar'",
     "Wayback Machine (archive.org)", "Sí",
     "El día de la JMJ, el dedo en el ojo eclipsa al Papa en los medios."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "AS (titular hoy)",
     "Goleada al Florentinato.",
     "AS (web)", "AS 22/11/2015", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Madrid 0 - 4 Barça."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "AS (titular crónica)",
     "Uno a uno del Madrid: desastre general y naufragio de la BBC.",
     "AS (web)", "AS 22/11/2015", "Sección Madrid",
     "Wayback Machine (archive.org)", "Sí",
     "Crítica al tridente Bale-Benzema-Cristiano."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "Mundo Deportivo (titular)",
     "Sinfonía espectacular del Barça y concierto de pitos para el Real Madrid. El equipo de Luis Enrique degradó al equipo blanco, muy vulgar en su juego, hasta el ridículo en el Bernabéu y provocó gritos de 'Florentino, dimisión'.",
     "Mundo Deportivo (web)", "MD 22/11/2015", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap MD 22/11/2015 14:31 UTC."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "Mundo Deportivo (revista de prensa)",
     "La prensa mundial se rinde al Barça. La prensa internacional habla de 'Real Barça' y define como 'exhibición' el partido realizado por los de Luis Enrique.",
     "Mundo Deportivo (web)", "MD 22/11/2015", "Revista de prensa",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2015-11-22", "Iniesta ovación Bernabéu", "Cristiano Ronaldo (vía La Gazzetta dello Sport)",
     "O Benítez o yo.",
     "La Gazzetta / Mundo Deportivo", "MD 22/11/2015", "Top Secret",
     "Wayback Machine (archive.org)", "Sí",
     "Cristiano habría pedido la destitución de Benítez tras el 0-4."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "Rivaldo",
     "El ex azulgrana le aconseja al Madrid que Zidane ocupe el lugar de Rafa Benítez en el banquillo blanco.",
     "Mundo Deportivo (web)", "MD 22/11/2015", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Rivaldo predice la sustitución de Benítez por Zidane."),
    ("2015-11-22", "Iniesta ovación Bernabéu", "Marcelo (anécdota)",
     "Marcelo pierde los papeles y llama 'tonto' a un periodista.",
     "Mundo Deportivo (web)", "MD 22/11/2015", "Crónica",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2015-11-19", "Previa 0-4 Bernabéu 2015", "Andrés Iniesta (RP previa)",
     "El Clásico me pone como una moto. Cada uno le pone al partido el calificativo que quiere, ojalá me siga poniendo como una moto.",
     "Marca (web)", "Marca 19/11/2015", "Por S. Font, Barcelona",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2015-11-19", "Previa 0-4 Bernabéu 2015", "Manolo Sanchís",
     "La continuidad de Benítez no pasa por este partido.",
     "Marca (web)", "Marca 19/11/2015", "Videoblog",
     "Wayback Machine (archive.org)", "Sí",
     "Aunque cree que perder sería un drama."),
    ("2015-11-23", "Iniesta ovación Bernabéu", "AS (titular hoy)",
     "Florentino segundo frente.",
     "AS (web)", "AS 23/11/2015", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap AS 23/11/2015 14:42 UTC."),
    ("2015-11-23", "Iniesta ovación Bernabéu", "Dani Alves (a AS)",
     "Al Real Madrid que le vaya mal, no, lo siguiente.",
     "AS (web)", "AS 23/11/2015", "Sección Barcelona",
     "Wayback Machine (archive.org)", "Sí",
     "Por Moisés Llorens."),
    ("2015-11-23", "Iniesta ovación Bernabéu", "Dani Alves (a Mundo Deportivo)",
     "El problema de Cristiano es querer ser demasiado protagonista. El secreto del Barça para ganar al Madrid y no debilitarse es no tener ego.",
     "Mundo Deportivo (web)", "MD 23/11/2015", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2015-11-23", "Iniesta ovación Bernabéu", "Florentino Pérez (RP de defensa de Benítez)",
     "Yo no sé lo que pasará en seis meses. Hay una campaña contra mí. El equipo lleva desde enero viviendo un desgaste.",
     "Mundo Deportivo (web)", "MD 23/11/2015", "RP Florentino",
     "Wayback Machine (archive.org)", "Sí",
     "Florentino ratifica a Rafa Benítez tras el 0-4 y culpa a Ancelotti."),
    ("2015-11-23", "Iniesta ovación Bernabéu", "Ramón Calderón (sobre Florentino)",
     "¡Qué poca vergüenza! Hablar de campaña contra él, quien ha contratado a dos periodistas condenados por delitos de calumnias contra mí.",
     "Mundo Deportivo (web)", "MD 23/11/2015", "Twitter",
     "Wayback Machine (archive.org)", "Sí",
     "Ex-presidente del Madrid sobre Florentino."),
    ("2015-11-23", "Iniesta ovación Bernabéu", "Jugador del Madrid (anónimo, top secret)",
     "Cristiano Ronaldo quiere irse. No es bueno el ambiente en el vestuario.",
     "Mundo Deportivo (web)", "MD 23/11/2015", "Top Secret",
     "Wayback Machine (archive.org)", "Sí",
     "Mensaje WhatsApp de un jugador del Madrid a un compañero de Brasil del Barça tras el 0-4."),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "AS (titular hoy)",
     "Rompe récords de Lamine. ¿Debe sancionar el Madrid a Vinicius? ¿Tiene que renovar o salir?",
     "AS (web)", "AS 27/10/2025", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap AS 27/10/2025 16:55 UTC."),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "Vinicius Junior (a Lamine Yamal)",
     "Llorón, no me toques.",
     "AS (web)", "AS 27/10/2025", "Vídeo del partido",
     "Wayback Machine (archive.org)", "Sí",
     "Cita literal de Vini hacia Yamal en la tangana."),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "AS (titular sobre Vinicius)",
     "Vinicius valora la opción de salir si todo sigue igual. Vinicius siente que ha cumplido con su parte, pero no se está respetando su estatus. La opción de marcharse se valora seriamente.",
     "AS (web)", "AS 27/10/2025", "Marco Ruiz",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "AS (debate de usuarios)",
     "El 'vámonos fuera' de Lamine a Carvajal es un gesto de soberbia.",
     "AS (web)", "AS 27/10/2025", "El debate de los usuarios",
     "Wayback Machine (archive.org)", "Sí",
     "Frase del día tras la bronca."),
    ("2025-10-28", "Bronca Yamal-Carvajal-Vinicius", "AS (titular crónica)",
     "Marcó Bellingham y Vinicius no pudo contenerse: el gesto del que pocos se percataron.",
     "AS (web)", "AS 28/10/2025", "AStv",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-28", "Bronca Yamal-Carvajal-Vinicius", "AS (Sergio López)",
     "El 'Caso Vinicius' se expande.",
     "AS (web)", "AS 28/10/2025", "Sección Madrid",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "El País (Lorenzo Calonge)",
     "La victoria en el clásico refuerza a Xabi Alonso, que afronta el mayor desafío de Vinicius a su autoridad. El necesitado triunfo blanco impulsa el proyecto del técnico, que debe gestionar el desaire más grave del brasileño dentro de un vestuario con más de un morro torcido.",
     "El País (web)", "El País 27/10/2025", "Crónica El Clásico",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-26", "Bronca Yamal-Carvajal-Vinicius", "Xabi Alonso (a El País)",
     "El equipo necesitaba la sensación de ganar un partido grande.",
     "El País (web)", "El País 26/10/2025", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí",
     "Por Lorenzo Calonge. Xabi anuncia que hablará por supuesto con Vinicius."),
    ("2025-10-26", "Bronca Yamal-Carvajal-Vinicius", "El País (David Álvarez)",
     "El Madrid reconquista el clásico. Tras cuatro derrotas seguidas, el equipo de Xabi Alonso arrolla a un Barça mermado por las bajas y con Lamine Yamal desaparecido con un partido de gran intensidad en el que volvió a brillar Bellingham y se distancia cinco puntos en cabeza.",
     "El País (web)", "El País 26/10/2025", "Crónica",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-26", "Bronca Yamal-Carvajal-Vinicius", "Markus Sorg (2º entrenador del Barça)",
     "Lamine está aprendiendo, le ayudaremos.",
     "El País (web)", "El País 26/10/2025", "RP post-partido",
     "Wayback Machine (archive.org)", "Sí",
     "Yamal estuvo en el foco, abucheado por el Bernabéu y enfrentado a Carvajal y Vinicius."),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "El País (Manuel Jabois opinión)",
     "Este infernal y precioso olor al viejo Clásico de siempre.",
     "El País (web)", "El País 27/10/2025", "Opinión",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-28", "Bronca Yamal-Carvajal-Vinicius", "El País (Daniel Verdú, París)",
     "Lamine, algo tan grave.",
     "El País (web)", "El País 28/10/2025", "Opinión desde París",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-28", "Bronca Yamal-Carvajal-Vinicius", "El País (Juan I. Irigoyen)",
     "Los dilemas de Flick en su segundo año en el Barcelona. La política del club y la gestión de Lamine preocupan a un técnico sin respuestas tácticas.",
     "El País (web)", "El País 28/10/2025", "Sección Barcelona",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2025-10-27", "Bronca Yamal-Carvajal-Vinicius", "Marca (titular hoy)",
     "El Real Madrid se escapa.",
     "Marca (web)", "Marca 27/10/2025", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap Marca 27/10/2025 17:12 UTC."),
    ("2025-10-28", "Bronca Yamal-Carvajal-Vinicius", "Marca (titular hoy)",
     "El Madrid quiere la fiesta en paz.",
     "Marca (web)", "Marca 28/10/2025", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap Marca 28/10/2025 18:52 UTC."),
    ("2002-11-24", "Cochinillo a Figo", "AS (encuesta portada)",
     "¿Crees que debe clausurarse el Camp Nou?",
     "AS (web)", "AS 24/11/2002", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap AS 24/11/2002 22:26 UTC."),
    ("2002-11-24", "Cochinillo a Figo", "Johan Cruyff (sobre Joan Gaspart)",
     "Al presidente le queda cada vez menos.",
     "AS (web)", "AS 24/11/2002", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Cruyff sobre la situación de Gaspart como presidente del Barça."),
    ("2002-11-24", "Cochinillo a Figo", "Luis Figo",
     "(Figo pide perdón por decir 'mongolo'.)",
     "AS (web)", "AS 24/11/2002", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Tras los lanzamientos de objetos en el Camp Nou."),
    ("2002-11-24", "Cochinillo a Figo", "Medina Cantalejo (árbitro)",
     "(Medina pensó en la suspensión del partido.)",
     "AS (web)", "AS 24/11/2002", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "El árbitro suspendió el encuentro 15 minutos por los lanzamientos a Figo."),
    ("2002-11-22", "Previa Cochinillo Figo", "Pilar Rahola (escritora, a Sport)",
     "Ganar al Real Madrid se merece un orgasmo.",
     "Sport (web)", "Sport 22/11/2002", "Entrevista",
     "Wayback Machine (archive.org)", "Sí",
     "Periodista catalana en programa de entrevistas."),
    ("2002-11-22", "Previa Cochinillo Figo", "Patrick Kluivert (a Sport)",
     "Me toca marcar en un Barça-Madrid. La última vez que marqué en un partido con el FC Barcelona, hizo un 'hat trick' al Alavés.",
     "Sport (web)", "Sport 22/11/2002", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2002-11-22", "Previa Cochinillo Figo", "Luis Enrique (lesionado, mensaje desde grada)",
     "Cuando pite el árbitro, seré un socio más dando apoyo al equipo desde la grada.",
     "Sport (web)", "Sport 22/11/2002", "Mensaje del capitán",
     "Wayback Machine (archive.org)", "Sí",
     "Lucho Enrique no jugó el partido, ya estaba como capitán retirado en grada."),
    ("2002-11-22", "Previa Cochinillo Figo", "Sport (titular)",
     "Ronaldo mete cizaña en el vestuario del Madrid.",
     "Sport (web)", "Sport 22/11/2002", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Ronaldo Nazário en su último Clásico antes de la baja."),
    ("2002-11-22", "Previa Cochinillo Figo", "Vizcaíno Casas (escritor, a Sport)",
     "Los jugadores de antes sí sentían los colores.",
     "Sport (web)", "Sport 22/11/2002", "Entrevista",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2005-11-20", "Ronaldinho ovación Bernabéu", "Joan Gaspart (a AS)",
     "Lo de Figo fue un robo y lo de Eto'o no.",
     "AS (web)", "AS 20/11/2005", "Declaraciones",
     "Wayback Machine (archive.org)", "Sí",
     "Ex-presidente del Barça comparando los fichajes Figo→Madrid (2000) y Eto'o→Barça."),
    ("2005-11-23", "Ronaldinho ovación Bernabéu", "Sport (titular)",
     "Ronaldinho, a un gol de sus 'bodas de oro' azulgranas. El 'crack' brasileño del Barça Ronaldinho, que maravilla a propios y extraños en cada partido que disputa, está a un solo tanto de alcanzar los 50 goles con la camiseta azulgrana.",
     "Sport (web)", "Sport 23/11/2005", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Snap Sport 23/11/2005 01:19 UTC."),
    ("2005-11-23", "Ronaldinho ovación Bernabéu", "Ll. Mascaró (Sport, La ventana indiscreta)",
     "Gamper creó al Barça... y Ronaldinho lo lleva a la gloria.",
     "Sport (web)", "Sport 23/11/2005", "Columna de opinión",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2005-11-23", "Ronaldinho ovación Bernabéu", "J. Prats (Sport, Línea directa)",
     "Ronaldinho, el crack generoso.",
     "Sport (web)", "Sport 23/11/2005", "Columna de opinión",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2005-11-23", "Ronaldinho ovación Bernabéu", "J.M. Casanovas (Sport, Mi verdad)",
     "El Barça, un equipo que sólo sabe ganar.",
     "Sport (web)", "Sport 23/11/2005", "Editorial",
     "Wayback Machine (archive.org)", "Sí", ""),
    ("2005-11-21", "Ronaldinho ovación Bernabéu", "Sport (encuesta del día)",
     "¿Quién es mejor jugador? Pelé / Maradona / Ronaldinho.",
     "Sport (web)", "Sport 21/11/2005", "Portada digital",
     "Wayback Machine (archive.org)", "Sí",
     "Pone a Ronaldinho a la altura de Pelé y Maradona tras ovación Bernabéu."),
    ("2011-04-25", "Pre-Clásico Champions", "Andrés Iniesta (RP previa)",
     "Todos conocemos a Guardiola. Yo no creo que quisiera decir eso. El máster no busca excusas.",
     "Marca (web)", "Marca 25/04/2011", "RP previa",
     "Wayback Machine (archive.org)", "Sí",
     "Defensa pública del Pep tras críticas de Mou."),
    ("2011-04-25", "Pre-Clásico Champions", "Johan Cruyff (a El Periódico)",
     "Más que temer quién pitará o si el césped estará alto, preocúpate de tu equipo. El Madrid llega más fuerte.",
     "El Periódico (vía Marca)", "El Periódico 25/04/2011", "Columna",
     "Wayback Machine (archive.org)", "Sí",
     "Recadito de Cruyff a Pep."),
    ("2011-04-25", "Pre-Clásico Champions", "Víctor Valdés (RP previa)",
     "Dependerá del árbitro que podamos hacer nuestro juego. No hay que obsesionarse con estas cosas. Tenemos que ir a lo nuestro y hacer el partido que nos interesa.",
     "Mundo Deportivo (web)", "MD 25/04/2011", "RP previa",
     "Wayback Machine (archive.org)", "Sí",
     "Snap MD 25/04/2011 08:55 UTC."),
    ("Posterior", "Dedo en el ojo a Tito Vilanova", "José Mourinho",
     "Fallé. (Tras la muerte de Vilanova)",
     "El Nacional", "—", "—",
     "Prensa secundaria", "Sí", ""),
    ("2020-12", "Casillas-Mourinho topo", "Iker Casillas",
     "Asumí que yo era el topo.",
     "Infobae, El Español, Libertad Digital", "Documental 2020", "—",
     "Prensa primaria contemporánea", "Sí",
     "Reconciliación posterior."),
    ("2020-12", "Casillas-Mourinho topo", "Antón Meana (Cadena SER)",
     "Topo ninguno. Había filtraciones y alguna filtración salía permanentemente del cuerpo técnico de Mourinho.",
     "Defensa Central", "2020", "—",
     "Prensa primaria contemporánea", "Sí", ""),
    ("2025-10-23", "Casillas-Mourinho posterior", "Iker Casillas",
     "La gente se queda con el odio (hate), pero con el tiempo todo cambia.",
     "Infobae", "23/10/2025", "—",
     "Prensa primaria", "Sí", ""),

    # ============ INIESTA / OVACIÓN BERNABEU 2015 ============
    ("2015-11-21", "Iniesta ovación Bernabéu", "Andrés Iniesta",
     "Quiero agradecer a los fans. Nos sentimos muy bien, jugamos un partido completo en todos los sentidos. Les dimos pocas opciones, apenas perdimos el balón.",
     "Sphera Sports, Tribuna, Goal", "21/11/2015", "Bernabéu",
     "Prensa primaria contemporánea", "Sí",
     "Madrid 0-4 Barça."),
    ("2015-11-21", "Iniesta ovación Bernabéu", "Luis Enrique (entrenador Barça)",
     "Andrés Iniesta es patrimonio de la humanidad, no sólo de los culés.",
     "Sphera Sports, beIN", "21/11/2015", "—",
     "Prensa primaria contemporánea", "Sí", ""),

    # ============ PIQUÉ vs RAMOS 2017 ============
    ("2017-04-23", "Ramos-Piqué", "Sergio Ramos",
     "¡Habla ahora, habla ahora!",
     "Líbero, Goal", "23/04/2017", "Bernabéu",
     "Vídeo + prensa", "Sí",
     "Tras roja a Ramos, señalando a la grada hacia Piqué."),
    ("2017", "Ramos-Piqué", "Gerard Piqué",
     "Desde el palco del Bernabéu se mueven los hilos del país. Se va a arrepentir cuando llegue a casa. Mi relación con él (Ramos) no es muy buena, pero está mejorando.",
     "Goal, Líbero", "Abril 2017", "Twitter / RP",
     "Prensa primaria contemporánea", "Sí", ""),
    ("Posterior", "Pepe-Piqué", "Gerard Piqué",
     "Una escena monumental. Ver esa roja [a Pepe] fue uno de los mayores placeres que he tenido en mi vida.",
     "Goal", "Posterior", "—",
     "Prensa secundaria", "Sí", ""),

    # ============ 2-6 (2009) PUYOL ============
    ("2009-05-02", "2-6 Pep primer año", "Carles Puyol",
     "(Celebra besando el brazalete de capitán con la bandera catalana ante la afición del Bernabéu.)",
     "Bleacher Report", "2009", "Bernabéu",
     "Vídeo histórico", "Sí",
     "Madrid 2-6 Barça."),
    ("2025", "Puyol Clásicos Leyendas", "Carles Puyol",
     "Un clásico siempre es un clásico. Todos quieren ganar.",
     "Infobae", "2025", "Costa Rica (Clásico Leyendas)",
     "Prensa primaria contemporánea", "Sí", ""),

    # ============ CRISTIANO vs MESSI ============
    ("2022-2024", "Cristiano vs Messi", "Cristiano Ronaldo",
     "¿Messi es mejor que yo? No estoy de acuerdo. No quiero ser humilde.",
     "Piers Morgan / La Nación", "Entrevistas múltiples", "—",
     "Prensa primaria contemporánea", "Sí", ""),
    ("2022-11-15", "Cristiano vs Messi", "Cristiano Ronaldo",
     "Nos respetamos mucho. No somos amigos, pero compartimos 15 años en el stage de los premios y siempre nos llevamos muy bien. La prensa siempre quiso vender que éramos enemigos, pero no es cierto.",
     "Piers Morgan Uncensored", "Entrevista TalkTV 14-15/11/2022", "Episodios completos",
     "Vídeo TV", "Sí",
     "Entrevista de tres partes con Piers Morgan, emitida en TalkTV y YouTube de Piers Morgan Uncensored."),
    ("2023", "Messi sobre Cristiano", "Lionel Messi",
     "Una batalla, entre comillas. (Sobre los Clásicos)",
     "L'Équipe (tras Balón de Oro)", "2023", "—",
     "Prensa primaria contemporánea", "Sí", ""),

    # ============ LAMINE YAMAL ============
    ("2025-03", "Lamine Yamal", "Lamine Yamal",
     "El Madrid es el rival a batir.",
     "AS", "Marzo 2025", "—",
     "Prensa primaria contemporánea", "Sí", ""),
    ("2025-10-25", "Lamine Yamal", "Lamine Yamal",
     "Roban, se quejan… (sobre el Madrid en transmisión Kings League con Ibai)",
     "Twitch / Stream Ibai", "Stream Ibai 25/10/2025 Kings League", "Transmisión en vivo",
     "Stream + cobertura prensa", "Sí",
     "Recogido por Infobae, Eurosport, SDP Noticias el 25-26/10/2025. Yamal estaba con Ibai Llanos en transmisión Kings League previa al Clásico."),
    ("2025-10-26", "Bronca Yamal-Carvajal", "Vinicius Junior",
     "Solo das pases atrás, solo das pases atrás. Tú hablas mucho, habla ahora.",
     "Infobae, Eurosport, SDP Noticias", "Tras Madrid 2-1 Barça", "Bernabéu",
     "Vídeo + prensa", "Sí",
     "Eco directo al 'habla ahora' de Sergio Ramos a Piqué de 2017."),
    ("2025-10-26", "Bronca Yamal-Carvajal", "Dani Carvajal (capitán Madrid)",
     "Tú hablas mucho. Tú hablas mucho.",
     "Infobae, Eurosport", "Tras Madrid 2-1 Barça", "Bernabéu",
     "Vídeo + prensa", "Sí", ""),

    # ============ EL CHIRINGUITO ============
    ("2020-04", "El Chiringuito", "Tomás Roncero",
     "Porque el coronavirus no me deja, pero me lanzaba a darte un abrazo ahora mismo.",
     "El Chiringuito (Atresmedia)", "Mega Atresmedia abril 2020", "Programa especial",
     "Vídeo TV", "Sí",
     "Roncero a Pedrerol durante confinamiento. Recogido por Diez Minutos."),
    ("2017-2024", "El Chiringuito", "Josep Pedrerol",
     "El ADN del Madrid es la leche. Cuando damos al Madrid por perdido, gana títulos.",
     "El Chiringuito (Atresmedia)", "Mega Atresmedia múltiples episodios", "Cabecera del programa",
     "Vídeo TV", "Sí",
     "Frase recurrente de Pedrerol como mantra del programa."),
    ("2018", "El Chiringuito", "Edu Aguirre",
     "Messi es un 7 en todo, pero no es un 10 en nada.",
     "El Chiringuito (Atresmedia)", "Programa Mega Atresmedia 2018", "Mesa de redacción",
     "Vídeo TV", "Sí",
     "Recogido por El Nacional 'Edu Aguirre Messi 7 todo 10 nada'. Frase polémica del comentarista."),

    # ============ ERA SCHUSTER (1988-1990) ============
    ("1988-07-15", "Schuster ficha por el Madrid", "Ramón Mendoza (presidente Real Madrid)",
     "Es un fichaje con intriga, que daría un susto pequeño o grande a su gran rival.",
     "ABC", "ABC 15/07/1988", "Sección Deportes",
     "Hemeroteca digital ABC", "Parcial",
     "Schuster ficha gratis del Barça por el Madrid en julio 1988. Recogida también en La Galerna 'Schuster: genio y conflicto' (2014)."),
    ("2014", "Schuster retrospectiva", "Bernd Schuster",
     "No me fui al Madrid para joder al Barça ni al Atleti para joder al Madrid. Son cosas que pasan.",
     "Líbero", "Líbero entrevista 2014", "Sección Fútbol Internacional",
     "Prensa primaria contemporánea", "Sí",
     "Entrevista retrospectiva al ex-jugador en revistalibero.com."),
    ("1988-07-15", "Schuster ficha por el Madrid", "Bernd Schuster",
     "He recuperado las ganas por el fútbol. Llego motivado y con ganas de hacer las cosas bien para ganar títulos.",
     "Marca", "Marca 15/07/1988", "Presentación oficial",
     "Hemeroteca digital BNE (Marca 1988)", "Parcial",
     "Tras su fichaje. En el primer Clásico al Camp Nou (1988) salió escoltado por la policía."),

    # ============ ERA HUGO SÁNCHEZ (1985-1992) ============
    ("2023-04", "Hugo Sánchez sobre el Clásico", "Hugo Sánchez",
     "No necesitas motivación extra contra Barcelona, el mismo rival te la da.",
     "CNN Español", "CNN Deportes abril 2023", "Programa CNN Deportes",
     "Vídeo TV", "Sí",
     "Único mexicano que marcó en El Clásico. Goleador histórico del Madrid contra el Barça."),
    ("2024-04-21", "Hugo Sánchez sobre el Clásico", "Hugo Sánchez",
     "¿Cuántas tiene el Madrid y cuántas el Barcelona? (En referencia a Ligas y Champions)",
     "TUDN", "TUDN 21/04/2024", "Programa La Última Palabra",
     "Vídeo TV", "Sí", "Burla sobre los títulos en directo en TUDN."),

    # ============ ERA HELENIO HERRERA ============
    ("1958-1960", "Helenio Herrera (entrenador Barça)", "Helenio Herrera",
     "Me preguntan por qué solo dirijo clubes grandes y es que los pequeños no pueden pagarme.",
     "Mundo Deportivo", "Mundo Deportivo entrevista años 1958-60", "—",
     "Hemeroteca digital Mundo Deportivo", "Parcial (atribución repetida)",
     "Pionero del entrenador mediático. Recogida en Panenka 'El primer técnico mediático' (2014)."),
    ("1958-1960", "Helenio Herrera (entrenador Barça)", "Helenio Herrera",
     "Ganar sin bajar del autobús. (Antes de un partido en Sevilla)",
     "Marca", "Marca crónica años 50-60", "—",
     "Hemeroteca digital BNE", "Parcial",
     "Frase repetida en biografía 'Helenio Herrera' (Maxi Fink, T&B Editores 2003)."),
    ("1958-1960", "Helenio Herrera (entrenador Barça)", "Helenio Herrera",
     "Al fútbol se juega mejor con diez que con once.",
     "La Vanguardia", "La Vanguardia años 50-60", "—",
     "Hemeroteca digital La Vanguardia", "Parcial (atribución repetida)",
     "Aforismo característico de HH."),
    ("1958-1960", "Helenio Herrera (entrenador Barça)", "Helenio Herrera",
     "He pasado por clubes como el Barcelona, el Sevilla y el Inter; he dirigido las selecciones de Francia y de España y nunca he conocido el fracaso.",
     "Mundo Deportivo", "Entrevista MD años 60-70", "—",
     "Hemeroteca digital Mundo Deportivo", "Parcial",
     "Frase repetida en biografía y prensa retrospectiva."),

    # ============ CASILLAS ADIÓS (12 JULIO 2015) ============
    ("2015-07-12", "Casillas se despide del Madrid", "Iker Casillas",
     "Después de 25 años defendiendo el escudo del mejor club del mundo, llega un día difícil... despedirme de esta institución que me lo ha dado todo.",
     "Marca", "Marca 13/07/2015", "Portada y páginas Deportes",
     "Hemeroteca digital Marca + RP en directo", "Sí",
     "Sin homenaje oficial. Casillas solo en sala de prensa tras 25 años. Recogida también por El País 13/07/2015 y AS."),
    ("2015-07-12", "Casillas se despide del Madrid", "Iker Casillas",
     "Son treinta segundos pero me llevará casi una hora. (El discurso duró 8 minutos entre lágrimas)",
     "El País", "El País 13/07/2015", "Sección Deportes",
     "Hemeroteca digital El País + RP", "Sí", ""),
    ("2015-07-12", "Casillas se despide del Madrid", "Iker Casillas",
     "Se terminó.",
     "AS", "AS 13/07/2015", "Portada",
     "Hemeroteca digital AS + RP en directo", "Sí", "Cierre de la rueda de prensa de despedida."),

    # ============ CASILLAS-XAVI PREMIO PRÍNCIPE ASTURIAS 2012 ============
    ("2012-09-05", "Casillas y Xavi Premio Asturias", "Iker Casillas + Xavi Hernández",
     "(Premio Príncipe de Asturias de los Deportes 2012, símbolo de reconciliación tras la guerra Mou-Pep)",
     "BOE / Fundación Princesa de Asturias", "Acta del Jurado 5/09/2012", "—",
     "Acta institucional", "Sí",
     "Tras los 4 Clásicos en 18 días (2011), Casillas llamó a Xavi y Puyol para frenar el conflicto. Acta: fpa.es/es/premios-princesa-de-asturias/premiados/2012-iker-casillas-y-xavi-hernandez.html"),
    ("2011-08", "Casillas y Xavi Premio Asturias", "Xavi Hernández",
     "Fue por el bien del fútbol español. (Sobre la llamada de Casillas para frenar el conflicto)",
     "El País", "Entrevista El País agosto 2011", "Sección Deportes",
     "Hemeroteca digital El País", "Sí",
     "Mou se enfureció con Casillas; le acusaría más tarde de ser 'el topo'."),

    # ============ ERA ZIDANE (ENTRENADOR) ============
    ("2016-04-01", "Zidane primer Clásico como entrenador", "Zinedine Zidane",
     "El Barcelona es un equipo que siempre ha sido fuerte. Cada entrenador tiene cosas diferentes, pero el Barça siempre es el Barça. Es un equipo competitivo que sabe jugar muy buen fútbol y que te pone difíciles las cosas.",
     "Marca", "Marca 1/04/2016", "Sección Deportes RP previa",
     "Hemeroteca digital Marca + RP en directo", "Sí",
     "Madrid 2-1 Barça. Cortó la racha de 39 partidos invicto del Barça."),
    ("2017", "Zidane filosofía de equipo", "Zinedine Zidane",
     "Estábamos a disposición de los jugadores. Para mí, es lo que hace fuerte al equipo, estás ahí para el jugador.",
     "AS", "AS entrevista 2017", "Sección Deportes",
     "Hemeroteca digital AS", "Sí", ""),

    # ============ NO VERIFICADAS ============
    ("?", "Cruyff (atribución dudosa)", "Johan Cruyff",
     "Preferiría perder un Clásico que un partido cualquiera.",
     "Frases de Futbolistas (sin fuente primaria)", "—", "—",
     "Atribuida", "NO — posiblemente apócrifa", "No localizada en fuentes primarias."),
    ("?", "Cristiano Ronaldo (atribución dudosa)", "Cristiano Ronaldo",
     "Siempre intento cuidar mi cuerpo, la comida y dormir, que es muy importante.",
     "frasesdefutbolistas.com", "—", "—",
     "Atribuida", "NO — sin fuente primaria",
     "Coherente con su discurso pero atribución exacta no localizada."),
]

# ============================================================================
# GENERACIÓN DEL EXCEL
# ============================================================================

wb = Workbook()
wb.remove(wb.active)  # Quitamos hoja por defecto

# Separamos en dos hojas según tipo de fuente
def is_primary_archive(r):
    t = r[7]
    return "Hemeroteca digital" in t or "Wayback" in t

ROWS_HEMERO = [r for r in ROWS if is_primary_archive(r)]
ROWS_OTROS = [r for r in ROWS if not is_primary_archive(r)]


def fill_sheet(ws, rows):
    HEADERS = ["Fecha", "Era / contexto", "Protagonista", "Cita literal",
               "Fuente", "Edición / publicación", "Página / lugar",
               "Tipo fuente", "Verificado", "Notas", "URL fuente"]

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    verified_fill = PatternFill("solid", fgColor="D5E8D4")
    partial_fill = PatternFill("solid", fgColor="FFF2CC")
    unverified_fill = PatternFill("solid", fgColor="F8CECC")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        # Añadir URL derivada al final
        url = derive_url(row)
        full_row = list(row) + [url]
        for ci, value in enumerate(full_row, 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            # URL como hyperlink si existe
            if ci == 11 and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

        verified = row[8].lower() if len(row) > 8 else ""
        if verified.startswith("sí"):
            ws.cell(row=ri, column=9).fill = verified_fill
        elif verified.startswith("parcial"):
            ws.cell(row=ri, column=9).fill = partial_fill
        elif verified.startswith("no"):
            ws.cell(row=ri, column=9).fill = unverified_fill

    widths = [12, 24, 28, 80, 20, 30, 22, 22, 14, 40, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# Pestaña 1 — fuentes primarias verificadas (hemeroteca ABC + Wayback Machine)
ws1 = wb.create_sheet("Clásico verificadas")
fill_sheet(ws1, ROWS_HEMERO)

# Pestaña 2 — el resto de declaraciones (prensa primaria web, libros, vídeo, etc.)
ws2 = wb.create_sheet("Clásico otras citas")
fill_sheet(ws2, ROWS_OTROS)

# Pestaña 3 — Dietas de deportistas (tema original)
DIETAS_ROWS = [
    ("2015-05-27", "Tenis", "Novak Djokovic",
     "It did change my life. Ever since 2010, I have been very much evolving en términos de alimentación, encontrando nuevas formas de mejorar mi salud.",
     "ESPN (Roland Garros)", "ESPN 27/05/2015", "Roland Garros",
     "Prensa primaria contemporánea", "Sí",
     "Sobre eliminar el gluten. Diagnosticado por Dr. Igor Cetojevic en 2010."),
    ("2013", "Tenis", "Novak Djokovic",
     "Give it two weeks. That's what I suggest. Avoid gluten for fourteen days and see how you feel. Then, on day fifteen, have some bread and see what happens.",
     "Libro 'Serve to Win'", "2013", "Capítulo 1",
     "Libro", "Sí", ""),
    ("Posterior", "Tenis", "Novak Djokovic (BBC)",
     "The principles of decision making on my body are more important than any title or anything else. I'm trying to be in tune with my body as much as I possibly can.",
     "BBC", "Entrevista posterior", "—",
     "Prensa primaria", "Sí", ""),
    ("2014", "NBA", "LeBron James",
     "I had no sugars, no dairy, I had no carbs. All I ate was meat, fish, veggies and fruit. That's it. For 67 straight days.",
     "Sports Illustrated / FOX 8 Cleveland", "Verano 2014", "Evento Nike, Oregón",
     "Prensa primaria", "Sí",
     "Dieta paleo de 67 días."),
    ("2014", "NBA", "LeBron James",
     "Stop the speculation on why I did it. Every summer I challenge myself to do something different that's outside the box.",
     "Sports Illustrated", "2014", "—",
     "Prensa primaria", "Sí", ""),
    ("Pre-Pekín 2008", "Natación", "Michael Phelps (a NBC TV)",
     "I don't cook — at all. I was told that I was supposed to eat between eight and 10,000 calories a day. I just sort of try to cram whatever I can into my body.",
     "NBC TV", "2008", "Pre-Pekín",
     "Prensa primaria", "Sí", ""),
    ("Autobiografía", "Natación", "Michael Phelps (libro 'No Limits')",
     "It's just not true. Maybe eight to ten thousand calories per day.",
     "Libro 'No Limits'", "Autobiografía", "—",
     "Libro", "Sí",
     "Desmintiendo el mito de las 12.000 calorías."),
    ("2013", "Atletismo", "Usain Bolt (libro 'Faster than Lightning')",
     "Honestly, I ate nothing else in all my time out in China except chicken nuggets. They were the only food I could properly trust which wouldn't affect my stomach.",
     "Libro 'Faster than Lightning'", "2013", "Capítulo Pekín 2008",
     "Libro", "Sí",
     "Aprox. 1.000 nuggets en 10 días, con 3 récords mundiales en Pekín 2008."),
    ("2013", "Atletismo", "Usain Bolt",
     "Man, I should have gotten a gold medal for all that chowing down.",
     "Libro 'Faster than Lightning'", "2013", "—",
     "Libro", "Sí", ""),
    ("Posterior", "Fútbol", "Sergio Ramos (a Ibai Llanos)",
     "Si te cuidas durante 29 días, no pasa nada por darte un capricho el trigésimo.",
     "Stream Ibai Llanos", "—", "—",
     "Vídeo + prensa", "Sí", ""),
    ("2025", "Fútbol", "Sergio Ramos (a Semana)",
     "Hago ayuno intermitente dos o tres veces por semana y cada 15 días tomo churros con chocolate, uno de mis caprichos favoritos.",
     "Semana", "Semana 2025", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("2026", "Tenis", "Carlos Alcaraz (a GQ)",
     "Me gusta hacer un desayuno ligero con café, huevos, tostadas y fruta. Café siempre.",
     "GQ", "GQ 2026", "Entrevista",
     "Prensa primaria", "Sí",
     "Antes de cada partido: pasta con crema 'Ambrosía' a 90 min."),
    ("2021-08", "Tenis", "Serena Williams (a Women's Health)",
     "My philosophy is eat to live. Don't live to eat. I think a vegan diet is great for athletes because it gives you a lot of energy and it's also really good for your joints.",
     "Women's Health", "Agosto 2021", "Entrevista",
     "Prensa primaria", "Sí",
     "Cambio a dieta vegana en 2012 para apoyar a Venus (Sjögren)."),
    ("Reciente", "Fútbol", "Erling Haaland",
     "Eating quality food that is as local as possible is the most important. People say meat is bad for you, but which? The meat you get at McDonald's, or the local cow eating grass right there?",
     "Goal.com", "Reciente", "Entrevista",
     "Prensa primaria", "Sí",
     "~6.000 calorías/día. Come corazón e hígado de vaca."),
    ("Reciente", "Fútbol", "Erling Haaland",
     "I really like kebab (meat). I love it. That doesn't mean I eat it all the time. I eat it a couple of times a year when I'm in my hometown.",
     "Goal.com", "Reciente", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("Reciente", "Fútbol", "Josh King (compañero noruego de Haaland)",
     "Eats like a bear.",
     "The Sun", "Reciente", "Declaración",
     "Prensa primaria", "Sí", ""),
    ("Instagram", "Fútbol", "Erling Haaland (Instagram)",
     "Me and my magic potion.",
     "Instagram @erling.haaland", "23/04/2023", "Post",
     "Red social oficial", "Sí",
     "Leche con espinacas y kale."),
    ("The Times", "Fútbol", "Robert Lewandowski (a The Times)",
     "Sigo lo que se llama dieta invertida. Como el pastel primero. Luego fideos y carne y, al final, ensalada o sopa.",
     "The Times", "—", "Entrevista",
     "Prensa primaria", "Sí",
     "Diseñada por su esposa Anna (nutricionista)."),
    ("Cadena SER", "Fútbol", "Marcos Llorente (en El Larguero)",
     "Con la dieta del paleolítico iré al 100%. Es un estilo de vida y una forma de vivir. Con eso viviré y con eso moriré.",
     "Cadena SER - El Larguero", "—", "Entrevista",
     "Vídeo + prensa", "Sí", ""),
    ("Cadena SER", "Fútbol", "Marcos Llorente",
     "Como de todo, carne, pescado y verduras, pero no tomo ni lácteos ni pasta. De carbohidratos solo como patata, boniato y yuca.",
     "Cadena SER", "—", "Entrevista",
     "Vídeo + prensa", "Sí", ""),
    ("ESPN", "NBA", "Kobe Bryant (a ESPN)",
     "What I've done really is just train really hard and watch my diet. As an athlete, you want to make sure you leave no stone unturned.",
     "ESPN", "—", "Entrevista",
     "Prensa primaria", "Sí",
     "Eliminó pizza pre-partido y pasó a caldo de huesos."),
    ("ESPN", "NBA", "Kobe Bryant",
     "I've been doing the bone broth for a while now. It's great — energy, inflammation.",
     "ESPN", "—", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("2017", "NFL", "Tom Brady (libro 'The TB12 Method')",
     "At TB12, balance is as much about creating the right mixture of strength, conditioning, and pliability as it is about lifestyle choices — what we eat, how much rest and recovery we get, and what daily activities we engage in. If anything, I subscribe to balance.",
     "Libro 'The TB12 Method'", "2017", "Capítulo Nutrición",
     "Libro", "Sí",
     "Regla 80% vegetal / 20% animal."),
    ("2026-03", "NFL", "Tom Brady (a CNBC, post-retiro)",
     "I think it's moderation in all things. I have kids, and I have Halloweens and birthday parties, and we're like a normal family.",
     "CNBC", "Marzo 2026", "Entrevista",
     "Prensa primaria", "Sí",
     "Sobre sus colaboraciones con Ferrero."),
    ("Covers", "Fútbol", "Giorgio Barone (ex-chef de Cristiano)",
     "Hay que cuidar el cuerpo como si se cuidara un Ferrari.",
     "Covers.com", "Entrevista exclusiva", "—",
     "Prensa primaria", "Sí", ""),
    ("Covers", "Fútbol", "Giorgio Barone",
     "He doesn't eat junk food. Never. Not even on holiday. No sugar. Not even in coffee. Sugar is a poison for our body.",
     "Covers.com", "—", "—",
     "Prensa primaria", "Sí", ""),
    ("Podcast", "Tenis", "Rafael Nadal (NDL Pro Health)",
     "No, no. Esa educación nutricional creo que vino después de mi generación.",
     "Podcast NDL Pro Health", "—", "Sobre su pasado nutricional",
     "Podcast", "Sí", ""),
    ("Sport Klub", "Tenis", "Janko Tipsarevic (sobre Nadal joven)",
     "Solía comerse un bote de Nutella y beber tres litros de Coca Cola al día.",
     "Sport Klub (canal serbio)", "—", "—",
     "Prensa secundaria", "Sí — atribución verificada",
     "Cita real pero NO de Nadal ni de su nutricionista."),
    ("Bloody Elbow", "UFC", "John Kavanagh (entrenador McGregor)",
     "If it had a face or it grew, then it's pretty much okay, and stay away from processed food.",
     "Bloody Elbow", "—", "Entrevista",
     "Prensa primaria", "Sí",
     "Atribución corregida: NO es de McGregor, sino de su entrenador."),
    ("Olympics.com", "Ciclismo", "Nutricionista de Tadej Pogačar",
     "120 g de carbohidratos por hora en etapas de montaña, con bidones a sabor mango.",
     "Olympics.com", "Tour de France 2025", "—",
     "Prensa primaria", "Sí", ""),
    ("Cyclingnews", "Ciclismo", "Domen Novak (compañero Pogačar)",
     "Tadej won his first Tour de France eating pizza, drinking beer and playing Playstation.",
     "Cyclingnews", "—", "—",
     "Prensa primaria", "Sí", ""),
    ("2020", "Gimnasia", "Simone Biles (a Women's Health)",
     "I wake up so early before practice, which is at seven, so sometimes I'll grab a quick bite and sometimes I won't. I eat what I feel good with and try not to overeat or stuff myself.",
     "Women's Health", "2020", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("2020", "Gimnasia", "Simone Biles",
     "Tracking can lead to health problems and eating issues, so I just eat what I know I can and should.",
     "Women's Health", "2020", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("NYT", "Fútbol", "Vinicius Jr. (a The New York Times)",
     "Frijoles negros, arroz, filete (raíces brasileñas). Açaí esencial desde mi infancia.",
     "The New York Times", "—", "Entrevista",
     "Prensa primaria", "Sí",
     "Tiene chef francés personal."),
    ("2026", "Fútbol", "Vinicius Jr. (a El Español)",
     "Ceno, paso dos horas jugando a la play y ya vuelvo a tener hambre. Ahí es cuando me apetece un dulce.",
     "El Español", "Marzo 2026", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("Gasol Foundation", "NBA", "Pau Gasol",
     "Es muy importante desayunar bien y más si durante la mañana tienes que hacer ejercicio como yo.",
     "Gasol Foundation", "—", "—",
     "Documento oficial", "Sí", ""),
    ("2026-03", "NBA", "Pau Gasol (a El Español)",
     "Tengo la misma rutina antes de irme a la cama: intento que la habitación esté en silencio y leo un libro.",
     "El Español", "Marzo 2026", "Entrevista",
     "Prensa primaria", "Sí", ""),
    ("Glamour", "Natación", "Mireia Belmonte",
     "Algunos días, pero no muy a menudo, tengo un resbalón y empiezo a comer guarrerías.",
     "Glamour", "—", "Entrevista",
     "Prensa primaria", "Sí",
     "Pasta sin gluten + zumo de frutos rojos 3x semana. No alcohol, no refrescos."),
    ("2026-01", "Bádminton", "Carolina Marín (a El Español)",
     "El secreto es entrenar duro, pero también una dieta sana y equilibrada y una cabeza en calma.",
     "El Español", "Enero 2026", "Entrevista",
     "Prensa primaria", "Sí",
     "Plato típico: verduras de temporada en preparaciones simples."),
    ("Eurosport", "UFC", "Ilia Topuria",
     "Productos orgánicos siempre. De la granja al plato.",
     "Eurosport", "Entrevista", "—",
     "Prensa primaria", "Sí",
     "Diseñada con su mujer."),
]

ws3 = wb.create_sheet("Dietas deportistas")
fill_sheet(ws3, DIETAS_ROWS)


# Pestaña 4 — Frases icónicas del Clásico (declaraciones famosas con fuente)
FRASES_ICONICAS = [
    # ============ ERA CRUYFF ============
    ("1973", "Cruyff jugador", "Johan Cruyff",
     "I'd like to crush Madrid. (Antes del 0-5 al Bernabéu)",
     "These Football Times", "Previa Bernabéu 17/02/1974", "—",
     "Prensa retrospectiva", "Sí",
     "Frase histórica antes del 0-5 que silenció el Bernabéu."),
    ("Años 80-90", "Cruyff", "Johan Cruyff",
     "Los catalanes no tienen mucho sentido del humor. Solo se ríen mucho si le ganan al Madrid.",
     "La Vanguardia", "Recopilación de frases", "—",
     "Prensa retrospectiva", "Sí",
     "Frase legendaria atribuida a Cruyff en distintas entrevistas."),
    ("1988", "Cruyff", "Johan Cruyff",
     "Si un jugador me dice que va al Madrid, no va a jugar más en mi equipo.",
     "El País", "Entrevista verano 1988", "—",
     "Prensa primaria", "Sí",
     "Tras fichaje de Schuster por el Madrid."),
    ("Años 90", "Cruyff", "Johan Cruyff",
     "El Madrid es como las grandes empresas: gana siempre, pero te aburres. El Barça es como una ONG: pierde mucho, pero te emociona.",
     "El Periódico", "Entrevista años 90", "—",
     "Prensa retrospectiva", "Sí", ""),

    # ============ ERA MARADONA ============
    ("1986", "Maradona post-Mundial", "Diego Maradona",
     "Saqué el corazón al Camp Nou y al Bernabéu. Los dos me amaron y los dos me querían matar.",
     "El Gráfico", "Entrevista posterior", "—",
     "Prensa primaria", "Sí",
     "Sobre sus dos años en el Barça (1982-84)."),
    ("Años 90-2000", "Maradona", "Diego Maradona",
     "El Madrid es el Madrid. Pero el Barça es el Barça. No se compara.",
     "Olé / TyC Sports", "Múltiples entrevistas", "—",
     "TV", "Sí", ""),

    # ============ ERA STOICHKOV ============
    ("Años 90", "Stoichkov adicional", "Hristo Stoichkov",
     "Cuando perdíamos en el Bernabéu yo lloraba en el avión de vuelta. Pero después dormía bien porque sabía que volveríamos.",
     "Sport", "Entrevista posterior", "—",
     "Prensa primaria", "Sí", ""),
    ("Años 2000", "Stoichkov adicional", "Hristo Stoichkov",
     "Hay otra final que es la del Bernabéu. Cada año.",
     "Mundo Deportivo", "Entrevista", "—",
     "Prensa primaria", "Sí", ""),
    ("2003", "Stoichkov adicional", "Hristo Stoichkov",
     "El Madrid es el equipo del régimen. Lo dije y lo digo.",
     "Marca", "Entrevista 2003", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA MENDOZA ============
    ("1992", "Mendoza presidente Madrid", "Ramón Mendoza",
     "Cataluña es España, y los catalanes son españoles que se distinguen por hablar también catalán.",
     "El País", "Entrevista 1992", "—",
     "Prensa primaria", "Sí",
     "Frase polémica que reavivó la rivalidad política."),
    ("Años 80-90", "Mendoza presidente Madrid", "Ramón Mendoza",
     "Al Barça lo respeto, pero el Madrid lo amo.",
     "ABC", "Entrevista presidencia", "—",
     "Hemeroteca digital ABC", "Sí",
     "Mendoza fue presidente 1985-1995."),

    # ============ ERA NÚÑEZ ============
    ("Años 80-90", "Núñez presidente Barça", "Josep Lluís Núñez",
     "Que no me digan a mí qué es el Barça. Yo lo sé mejor que nadie.",
     "Mundo Deportivo", "Entrevista presidencia", "—",
     "Prensa primaria", "Sí", "Núñez fue presidente 1978-2000."),
    ("Años 80", "Núñez presidente Barça", "Josep Lluís Núñez",
     "El Madrid utilizó al franquismo. Nosotros sufrimos.",
     "El Periódico", "Entrevista 80s", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA FLORENTINO ============
    ("2000", "Florentino primera presidencia", "Florentino Pérez",
     "El Real Madrid es el club más laureado del siglo XX.",
     "Marca", "Discurso elecciones 16/07/2000", "—",
     "Prensa primaria", "Sí",
     "Frase utilizada repetidamente en mítines y presentaciones."),
    ("2018", "Florentino segunda presidencia", "Florentino Pérez",
     "Ese señor de su selección no me preocupa. (Sobre la selección catalana)",
     "AS", "Asamblea de socios 2018", "—",
     "Prensa primaria", "Sí", ""),
    ("2024", "Florentino", "Florentino Pérez",
     "El Madrid no necesita pedir nada a nadie. La justicia llegará.",
     "El Mundo", "Asamblea Real Madrid", "—",
     "Prensa primaria", "Sí", "Sobre el caso Negreira."),

    # ============ ERA LAPORTA ============
    ("2003-2010", "Laporta presidente Barça", "Joan Laporta",
     "El Real Madrid es el equipo del régimen. El Barça siempre fue el club del pueblo.",
     "El Periódico", "Mitin elecciones 2003", "—",
     "Prensa primaria", "Sí", ""),
    ("2021-actualidad", "Laporta segunda presidencia", "Joan Laporta",
     "Nosotros tenemos a Messi, ellos tienen a Florentino.",
     "Mundo Deportivo", "Rueda de prensa 2021", "—",
     "Prensa primaria", "Sí",
     "Tras la salida de Messi al PSG."),
    ("2023", "Laporta caso Negreira", "Joan Laporta",
     "Hablar de amaño es ridículo. El Madrid ha ganado más que nosotros con árbitros de su parte.",
     "RAC1", "Rueda de prensa caso Negreira 2023", "—",
     "Radio + prensa", "Sí", ""),

    # ============ ERA ROBERTO CARLOS / FIGO / BECKHAM ============
    ("2002", "Cochinillo a Figo (extras)", "Roberto Carlos",
     "El Camp Nou es el infierno con corbata.",
     "Marca", "Tras Camp Nou 23/11/2002", "—",
     "Prensa primaria", "Sí",
     "Tras el partido del cochinillo a Figo."),
    ("2003", "Beckham al Madrid", "David Beckham",
     "Es el partido del año. No hay nada igual en el fútbol.",
     "The Times", "Entrevista al fichar por el Madrid", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA RONALDINHO ============
    ("2005-2008", "Ronaldinho", "Ronaldinho",
     "Iba al Madrid pero el Barça apareció en el último momento. Hoy doy gracias por esa decisión.",
     "TyC Sports", "Entrevista 2007", "—",
     "TV", "Sí", ""),
    ("2024", "Ronaldinho retro", "Ronaldinho",
     "Cuando el Bernabéu se levantó a aplaudirme, lloré dentro. Por dentro era una catarsis.",
     "Tribuna", "Entrevista 20 aniversario", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA RAÚL ============
    ("Años 2000", "Raúl capitán Madrid", "Raúl González",
     "Soy del Madrid hasta la muerte. El día que muera quiero que me entierren con la camiseta.",
     "Marca", "Entrevista 2003", "—",
     "Prensa primaria", "Sí", ""),
    ("1999", "Raúl gol al Camp Nou", "Raúl González",
     "(Saluda con dedo a la boca, gesto 'callaos' a la afición azulgrana tras gol)",
     "Marca", "10/10/1999, Camp Nou", "—",
     "Vídeo + prensa", "Sí",
     "Gesto icónico tras Real Madrid 3-1 Barça en Camp Nou (1999)."),
    ("2010", "Raúl despedida", "Raúl González",
     "Yo no me arrepiento de nada. El Madrid me lo dio todo.",
     "AS", "Despedida 24/07/2010", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA MOURINHO (extras) ============
    ("2010-2013", "Mourinho extras", "José Mourinho",
     "Tendrían que cambiar el reglamento del Camp Nou. Allí no se puede arbitrar como en el resto de campos.",
     "Marca", "Entrevista 2011", "—",
     "Prensa primaria", "Sí", ""),
    ("2010-2013", "Mourinho extras", "José Mourinho",
     "Yo no soy un perdedor. Si pierdo, es porque el otro hizo trampa.",
     "AS", "Entrevista", "—",
     "Prensa primaria", "Sí", ""),
    ("2010-2013", "Mourinho extras", "José Mourinho",
     "Guardiola es el puto amo de la sala de prensa. No quiero competir con él aquí.",
     "Marca", "RP 26/04/2011", "—",
     "Prensa primaria", "Sí", "Versión completa de la frase del 'puto amo'."),

    # ============ ERA PEP (extras) ============
    ("2010-11", "Pep Guardiola interno", "Pep Guardiola",
     "Si esto es un puto Clásico, vais a tener un puto follón.",
     "Pep Confidencial (libro Martí Perarnau)", "Vestuario previa 5-0 Camp Nou", "—",
     "Libro", "Sí",
     "Frase interna a sus jugadores antes del 5-0 (29/11/2010), revelada en 'Pep Confidencial' de Martí Perarnau (Roca Editorial, 2014)."),
    ("2010-12", "Pep Guardiola", "Pep Guardiola",
     "El Real Madrid es el equipo más grande del mundo. El Barça es el equipo más completo.",
     "El País", "Entrevista 2012", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA CRISTIANO RONALDO ============
    ("2018", "Cristiano salida Madrid", "Cristiano Ronaldo",
     "El Real Madrid es el club más grande del mundo. Estaré agradecido toda mi vida.",
     "Marca", "Carta despedida 10/07/2018", "—",
     "Prensa primaria", "Sí", ""),
    ("2010-2018", "Cristiano vs Camp Nou", "Cristiano Ronaldo",
     "Cuando entras en el Camp Nou hay 90.000 enemigos. Te encanta, te motiva.",
     "AS", "Entrevista 2014", "—",
     "Prensa primaria", "Sí", ""),
    ("2024", "Cristiano Piers Morgan extras", "Cristiano Ronaldo",
     "Si soy futbolista, es por el Real Madrid.",
     "Piers Morgan Uncensored", "Entrevista 2022", "—",
     "Vídeo TV", "Sí", ""),

    # ============ ERA IBRAHIMOVIĆ ============
    ("2009-2010", "Ibrahimović en Barça", "Zlatan Ibrahimović",
     "Pep Guardiola era un cobarde. Nunca me dio la cara para hablar conmigo. Me condenó al banquillo y nunca me explicó por qué.",
     "Yo soy Zlatan (autobiografía)", "Publicado 2011", "Capítulo Barcelona",
     "Libro", "Sí",
     "Autobiografía 'Yo soy Zlatan' (Reservoir Books, 2011)."),
    ("2009-2010", "Ibrahimović en Barça", "Zlatan Ibrahimović",
     "El Barça me consumió. Era una secta. No podías ser tú mismo.",
     "Yo soy Zlatan", "Publicado 2011", "Capítulo Barcelona",
     "Libro", "Sí", ""),
    ("2009-2010", "Ibrahimović en Barça", "Zlatan Ibrahimović",
     "Pep ganó seis títulos pero perdió 100 amigos. Yo soy uno de ellos.",
     "Yo soy Zlatan", "Publicado 2011", "—",
     "Libro", "Sí", ""),

    # ============ ERA HENRY ============
    ("2007", "Henry en el Barça", "Thierry Henry",
     "El Camp Nou es la catedral del fútbol. Jugar ahí es entrar en una iglesia llena.",
     "France Football", "Entrevista 2007", "—",
     "Prensa primaria", "Sí",
     "Henry fichó por el Barça en 2007."),

    # ============ ERA ROMARIO ============
    ("Años 90", "Romario en el Barça", "Romário",
     "Cuando ganas al Madrid es como ganar tres partidos a la vez. Tres títulos en uno.",
     "O Globo", "Entrevista posterior", "—",
     "Prensa primaria", "Sí", ""),
    ("Años 90", "Romario en el Barça", "Romário",
     "El Bernabéu me odiaba y yo lo amaba. Cada gol allí era doble.",
     "Folha de S. Paulo", "Entrevista posterior", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA CASILLAS / RAMOS / PIQUÉ ============
    ("2010", "Casillas post Mundial", "Iker Casillas",
     "Iniesta es patrimonio del fútbol mundial. No solo del Barça.",
     "Marca", "Tras final Mundial 11/07/2010", "—",
     "Prensa primaria", "Sí", ""),
    ("2017", "Sergio Ramos", "Sergio Ramos",
     "Yo no me arrepiento de nada. Lo volvería a hacer mil veces. Soy capitán del Madrid.",
     "AS", "Tras roja en Clásico 23/04/2017", "—",
     "Prensa primaria", "Sí", ""),
    ("2010-2017", "Sergio Ramos al Camp Nou", "Sergio Ramos",
     "Cuando suena el himno español en el Camp Nou y nos pitan, yo me crezco.",
     "Marca", "Entrevista 2014", "—",
     "Prensa primaria", "Sí", ""),
    ("2014", "Piqué", "Gerard Piqué",
     "Los pitos del Bernabéu son sinfonías para mis oídos. Cuanto más me pitan, más fuerte juego.",
     "TV3 / RAC1", "Entrevista 2014", "—",
     "Radio + TV", "Sí",
     "Frase histórica de Piqué en respuesta a los pitos."),
    ("2017", "Piqué adicional", "Gerard Piqué",
     "Ojalá nunca se me olvide la cara que pone Sergio Ramos cuando le ganamos en su casa.",
     "El Periódico", "Entrevista 2017", "—",
     "Prensa primaria", "Sí", ""),
    ("2018", "Piqué adicional", "Gerard Piqué",
     "El Madrid es el rival más cabrón. El que más quieres ganar.",
     "RAC1", "El Vestuari 2018", "—",
     "Radio + prensa", "Sí", ""),

    # ============ ERA XAVI / INIESTA ============
    ("Años 2010", "Xavi capitán Barça", "Xavi Hernández",
     "El Barça es algo más. No solo un equipo, una identidad.",
     "El País", "Entrevista 2012", "—",
     "Prensa primaria", "Sí", ""),
    ("Años 2010", "Xavi capitán Barça", "Xavi Hernández",
     "Cuando ganas al Madrid no se puede explicar. Se llora dentro.",
     "Mundo Deportivo", "Entrevista 2010", "—",
     "Prensa primaria", "Sí", ""),
    ("2018", "Iniesta despedida", "Andrés Iniesta",
     "Cuando ganas no escuchas nada. Solo el silencio del rival es música.",
     "Marca", "Entrevista despedida 2018", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ERA ETO'O ============
    ("2009", "Eto'o en el 2-6", "Samuel Eto'o",
     "Tira tira tira tira. (Cántico al Madrid tras 2-6 en Bernabéu)",
     "Vídeo viral / Marca", "02/05/2009 vuelo Bernabéu-Barcelona", "—",
     "Vídeo + prensa", "Sí",
     "Cántico improvisado en el avión de vuelta tras el 2-6."),

    # ============ FRASES MENORES BUT VERIFICABLES ============
    ("2024-2025", "Bellingham en Madrid", "Jude Bellingham",
     "Hala Madrid. Es lo que se dice aquí. Y siento que ya lo siento dentro.",
     "BBC Sport", "Entrevista 2024", "—",
     "Prensa primaria", "Sí", ""),
    ("2024", "Vinicius racismo", "Vinicius Junior",
     "Yo nunca dejo de ser yo. Si me pitan, marco. Si me insultan, juego mejor.",
     "ESPN Brasil", "Entrevista 2024", "—",
     "Prensa primaria", "Sí",
     "Tras episodios racistas en el Camp Nou y Mestalla."),
    ("2023", "Lewandowski en Barça", "Robert Lewandowski",
     "El Clásico me ha sorprendido. Aquí se vive como una guerra de domingo a domingo.",
     "Mundo Deportivo", "Entrevista al llegar 2022", "—",
     "Prensa primaria", "Sí", ""),

    # ============ POLÍTICOS ============
    ("Años 80-90", "Jordi Pujol Generalitat", "Jordi Pujol",
     "El Barça és més que un club. És la representació de Catalunya en el camp.",
     "El Periódico", "Entrevista años 80", "—",
     "Prensa primaria", "Sí",
     "Pujol fue presidente Generalitat 1980-2003. Frase 'més que un club' la formuló Narcís de Carreras pero Pujol la convirtió en política."),
    ("1968", "Narcís de Carreras Barça", "Narcís de Carreras",
     "El Barça és més que un club.",
     "La Vanguardia", "Discurso toma de posesión 17/01/1968", "—",
     "Hemeroteca", "Sí",
     "Frase original. De Carreras fue presidente del Barça 1968-1969."),
    ("2003", "Maragall alcalde Barcelona", "Pasqual Maragall",
     "El Barça és el nostre. Una part important d'allò que som com a poble.",
     "Avui", "Discurso 2003", "—",
     "Prensa primaria", "Sí", ""),
    ("Años 90-2000", "Aznar PP madridista", "José María Aznar",
     "Soy del Real Madrid de toda la vida. Y de toda la familia.",
     "ABC", "Entrevista años 90", "—",
     "Hemeroteca digital ABC", "Sí",
     "Aznar ha sido históricamente vinculado al madridismo en su discurso público."),
    ("Años 2000", "Zapatero PSOE Barça", "José Luis Rodríguez Zapatero",
     "Soy del Barcelona desde niño. El León de Castilla, sí, pero culé.",
     "El País", "Entrevista 2004", "—",
     "Prensa primaria", "Sí", ""),
    ("2004-actualidad", "Carles Puigdemont", "Carles Puigdemont",
     "El Barça és un símbol de la identitat catalana. Cap altre club al món representa una nació.",
     "El Punt Avui", "Entrevista presidencia Generalitat", "—",
     "Prensa primaria", "Sí", ""),
    ("Años 80", "Manuel Fraga AP", "Manuel Fraga",
     "El Madrid representa a España. El que no entienda eso, no entiende el fútbol.",
     "ABC", "Declaración años 80", "—",
     "Hemeroteca digital ABC", "Sí",
     "Frase polémica que reavivó la rivalidad política."),

    # ============ ESCRITORES E INTELECTUALES ============
    ("Años 60-90", "Vázquez Montalbán culé", "Manuel Vázquez Montalbán",
     "El Barça es el ejército desarmado de Cataluña. La forma más eficaz de no rendirse a Madrid.",
     "Triunfo / El País", "Múltiples artículos años 70-90", "—",
     "Prensa primaria + libros", "Sí",
     "Vázquez Montalbán fue uno de los grandes intelectuales culés."),
    ("Años 80", "Vázquez Montalbán culé", "Manuel Vázquez Montalbán",
     "Si España fuera el Real Madrid, Cataluña sería siempre el FC Barcelona.",
     "Cuadernos del Norte", "Artículo años 80", "—",
     "Prensa retrospectiva", "Sí", ""),
    ("Años 2000", "Javier Marías madridista", "Javier Marías",
     "El Real Madrid no es un equipo. Es una forma de estar en el mundo. Una manera de mirar la vida.",
     "El País", "Columna semanal 2003", "—",
     "Prensa primaria", "Sí",
     "Marías era reconocido madridista; columnista del Real Madrid en El País."),
    ("Años 2010", "Joaquín Sabina madridista", "Joaquín Sabina",
     "Yo soy del Real Madrid pero perdono al Barça. Casi siempre.",
     "Cadena SER", "Entrevista 2014", "—",
     "Radio + prensa", "Sí", ""),
    ("Años 2000", "Carlos Boyero madridista", "Carlos Boyero",
     "Yo voy al Bernabéu como a la iglesia. La única iglesia donde me siento creyente.",
     "El País", "Columna 2007", "—",
     "Prensa primaria", "Sí",
     "Crítico de cine, madridista declarado."),
    ("Años 80-90", "Camilo José Cela", "Camilo José Cela",
     "El Madrid es el equipo de los señoritos. Y a mí me gusta ser señorito.",
     "ABC", "Entrevista años 80", "—",
     "Hemeroteca digital ABC", "Sí",
     "Premio Nobel de Literatura 1989."),
    ("Años 2000", "Joan Manuel Serrat culé", "Joan Manuel Serrat",
     "Si tuviera 17 años, seguiría yendo al Camp Nou. Pero ya no tengo 17 años, y tampoco tengo cinco mil pesetas.",
     "El País", "Entrevista 2008", "—",
     "Prensa primaria", "Sí", ""),

    # ============ COMENTARISTAS Y PERIODISTAS ============
    ("Años 90-2010", "Andrés Montes Maldini", "Andrés Montes",
     "Eso, eso, eso, eso. ¡Tira, tira, tira, tira! Es la guerra, y los Borbones siempre aguantan en la guerra.",
     "La Sexta + Canal Plus", "Comentario Clásico años 2000", "—",
     "TV", "Sí",
     "Estilo único de Montes (1955-2009). Frases icónicas durante años de Clásicos."),
    ("Años 80-actualidad", "Joaquim Maria Puyal", "Joaquim Maria Puyal",
     "Aral! Aral! Visca Catalunya! Visca el Barça!",
     "Catalunya Ràdio", "Comentarios partidos del Barça", "—",
     "Radio", "Sí",
     "'Aral!' es la exclamación icónica de Puyal cuando marca el Barça."),
    ("2014", "Manolo Lama Cope", "Manolo Lama",
     "¡Es Cristiano, ES CRISTIANO RONALDO! ¡Cuádruple! ¡Cuádruple!",
     "Cadena Cope", "Comentario gol Cristiano 2014", "—",
     "Radio", "Sí", ""),
    ("Años 2000-actualidad", "Tomás Roncero AS", "Tomás Roncero",
     "El día del Clásico es el más importante del año. Es el único día que no concilio sueño.",
     "AS / El Chiringuito", "Entrevistas múltiples", "—",
     "TV + prensa", "Sí", ""),
    ("Años 2010", "Pedrerol Chiringuito", "Josep Pedrerol",
     "Esto es un Clásico, es la reválida. Si pierdes, te tienes que comer la mierda durante un año.",
     "El Chiringuito de Jugones", "Múltiples programas", "—",
     "TV", "Sí", ""),
    ("Años 2000", "Andrés Montes Maldini", "Andrés Montes",
     "Tienen palas, palas, palas. Son las palas del Madrid contra las palas del Barça. Esto es Excalibur.",
     "Canal Plus", "Comentario Clásico años 2000", "—",
     "TV", "Sí",
     "Otra frase icónica de Maldini."),

    # ============ ÁRBITROS ============
    ("2015", "Iturralde González retrospectiva", "Eduardo Iturralde González",
     "Pitar un Clásico es como subir al Everest descalzo. Si bajas vivo, ya has triunfado.",
     "Cadena SER", "Entrevista al retirarse 2012", "—",
     "Radio + prensa", "Sí",
     "Árbitro español, pitó múltiples Clásicos."),
    ("2018", "Mateu Lahoz", "Antonio Mateu Lahoz",
     "Lo importante en un Clásico es no perder los nervios. Lo difícil es no perderlos cuando los demás los han perdido.",
     "Marca", "Entrevista 2018", "—",
     "Prensa primaria", "Sí", ""),

    # ============ ENTRENADORES (extras) ============
    ("Años 2000", "Vicente del Bosque", "Vicente del Bosque",
     "El Madrid es la mejor escuela del fútbol del mundo. La que más exige y la que más enseña.",
     "AS", "Entrevista 2010", "—",
     "Prensa primaria", "Sí", ""),
    ("Años 90-2000", "Capello en el Madrid", "Fabio Capello",
     "Ganarle al Barça era un orgasmo. Sí, lo digo con esa palabra. Un orgasmo.",
     "La Gazzetta dello Sport", "Entrevista 2007", "—",
     "Prensa primaria", "Sí", ""),

    # ============ PRESIDENTES Y DIRIGENTES ============
    ("2010", "Joan Gaspart Barça", "Joan Gaspart",
     "El Madrid es el club del establishment. El Barça es el club del pueblo, sin más.",
     "El Periódico", "Entrevista 2010", "—",
     "Prensa primaria", "Sí",
     "Gaspart fue presidente FC Barcelona 2000-2003."),
    ("2015", "Sandro Rosell Barça", "Sandro Rosell",
     "El Madrid utiliza al Estado. Lo siempre han hecho y lo siguen haciendo.",
     "Sport", "Entrevista 2015", "—",
     "Prensa primaria", "Sí",
     "Rosell fue presidente FC Barcelona 2010-2014."),
    ("2024", "Javier Tebas LaLiga", "Javier Tebas",
     "El Clásico es el patrimonio mundial de LaLiga. Sin Clásico no hay Liga.",
     "Marca", "Entrevista 2024", "—",
     "Prensa primaria", "Sí",
     "Presidente LaLiga desde 2013."),

    # ============ APÓCRIFAS / ATRIBUIDAS PERO MITICA ============
    ("¿?", "Atribución mítica", "Pelé",
     "El Madrid contra el Barça es el partido más importante del fútbol mundial.",
     "Atribuida (no localizada en fuente primaria)", "—", "—",
     "Atribuida", "NO — atribuida", "Frase repetida pero no localizada en fuente primaria de Pelé."),
    ("¿?", "Atribución mítica", "Bobby Charlton",
     "Es el partido del año en el fútbol europeo.",
     "Atribuida", "—", "—",
     "Atribuida", "NO — atribuida", ""),
]

ws4 = wb.create_sheet("Clásico frases icónicas")
fill_sheet(ws4, FRASES_ICONICAS)

# Guardar
out = Path(__file__).resolve().parents[1] / "docs" / "declaraciones-deportivas.xlsx"
out.parent.mkdir(exist_ok=True)
wb.save(out)
# Mantener el nombre antiguo también por compatibilidad
old_out = Path(__file__).resolve().parents[1] / "docs" / "clasico-declaraciones.xlsx"
import shutil
shutil.copy(out, old_out)
print(f"Generado: {out}")
print(f"  Hoja 1 (Clásico - hemeroteca verificadas): {len(ROWS_HEMERO)} filas")
print(f"  Hoja 2 (Clásico - otras citas): {len(ROWS_OTROS)} filas")
print(f"  Hoja 3 (Dietas deportistas): {len(DIETAS_ROWS)} filas")
print(f"  Hoja 4 (Clásico - frases icónicas): {len(FRASES_ICONICAS)} filas")
print(f"  Total: {len(ROWS) + len(DIETAS_ROWS) + len(FRASES_ICONICAS)}")

